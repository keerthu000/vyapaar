from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render,redirect
from django.contrib.auth.models import User, auth
from django.utils.text import capfirst
from django.contrib import messages
from . models import *
import json
from django.http.response import JsonResponse
from django.utils.crypto import get_random_string
from datetime import date
from datetime import timedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.template.response import TemplateResponse
from django.db.models import F
from openpyxl import load_workbook
from django.http.response import JsonResponse, HttpResponse
from openpyxl import Workbook
from num2words import num2words
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from io import BytesIO
from xhtml2pdf import pisa
from django.core.mail import send_mail, EmailMessage
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt


# Create your views here.
def home(request):
  return render(request, 'home.html')



    







# @login_required(login_url='login')
def logout(request):
    auth.logout(request)
    return redirect('/')

def view_profile(request):
  com =  company.objects.get(user = request.user) 
  selected_options = request.session.get('selected_options', None)
  
  context = {
              'company' : com,
              'selected_options': json.dumps(selected_options)
          }
  return render(request,'profile.html',context)
  
def edit_profile(request,pk):
  com= company.objects.get(id = pk)
  user1 = User.objects.get(id = com.user_id)
  selected_options = request.session.get('selected_options', None)

  if request.method == "POST":

      user1.first_name = capfirst(request.POST.get('f_name'))
      user1.last_name  = capfirst(request.POST.get('l_name'))
      user1.email = request.POST.get('email')
      com.contact_number = request.POST.get('cnum')
      com.address = capfirst(request.POST.get('ards'))
      com.company_name = request.POST.get('comp_name')
      com.company_email = request.POST.get('comp_email')
      com.city = request.POST.get('city')
      com.state = request.POST.get('state')
      com.country = request.POST.get('country')
      com.pincode = request.POST.get('pinc')
      com.gst_num = request.POST.get('gst')
      com.pan_num = request.POST.get('pan')
      com.business_name = request.POST.get('bname')
      com.company_type = request.POST.get('comp_type')
      if len(request.FILES)!=0 :
          com.profile_pic = request.FILES.get('file')

      com.save()
      user1.save()
      return redirect('view_profile')

  context = {
      'company' : com,
      'user1' : user1,
      'selected_options': json.dumps(selected_options)
  } 
  return render(request,'company/edit_profile.html',context)


def sale_invoices(request):
  return render(request, 'sale_invoices.html')

def estimate_quotation(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    all_estimates = Estimate.objects.filter(company = com)
    estimates = []
    for est in all_estimates:
      history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
      dict = {'estimate':est,'history':history}
      estimates.append(dict)
    context = {
      'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
    }
    return render(request, 'staff/estimate_quotation.html',context)

def payment_in(request):
  return render(request, 'payment_in.html')
    
def sale_order(request):
  return render(request, 'sale_order.html')

def delivery_chellan(request):
  return render(request, 'delivery_chellan.html')

def sale_return_cr(request):
  return render(request, 'sale_return_cr.html')


# created by athul
def settings(request):
  com =  company.objects.get(user = request.user)
  selected_options = request.session.get('selected_options', None)
  
  context = {
              'company' : com,
              'selected_options': json.dumps(selected_options),
              
          }
  return render(request, 'company/settings.html',context)

def hide_options(request):
    
    com =  company.objects.get(user = request.user)
    if request.method == 'POST':
        selected_options = list(request.POST.getlist('selected_options'))

    request.session['selected_options'] = selected_options
    
    context = {'selected_options': json.dumps(selected_options),
               'company' : com}
   
    return render(request, 'company/homepage.html', context)

# ------created by athul------

def company_reg(request):
  return render(request,'company/register.html')


    
    
def company_reg2(request):
  terms=payment_terms.objects.all()
  return render(request,'company/register2.html',{'terms':terms})  

def add_company(request):
  
  if request.method == 'POST':
    email=request.POST['email']
    user=User.objects.get(email=email)
    
    c =company.objects.get(user = user)
    c.company_name=request.POST['cname']

    c.address=request.POST['address']
    c.city=request.POST['city']
    c.state=request.POST['state']
    c.country=request.POST['country']
    c.pincode=request.POST['pincode']
    c.pan_number=request.POST['pannumber']
    c.gst_type=request.POST['gsttype']
    c.gst_no=request.POST['gstno']
    c.profile_pic=request.FILES.get('image')

    select=request.POST['select']
    terms=payment_terms.objects.get(id=select)
    c.dateperiod=terms
    c.start_date=date.today()
    days=int(terms.days)

    
    end= date.today() + timedelta(days=days)
    c.End_date=end

    code=get_random_string(length=6)
    if company.objects.filter(Company_code = code).exists():
       code2=get_random_string(length=6)
       c.Company_code=code2
    else:
      c.Company_code=code

   
    c.save()

    staff = staff_details.objects.get(email=email,position='company',company=c)
    staff.first_name = request.POST['cname']
    staff.last_name = ''
    staff.save()
    # messages.success(request, 'Welcome'+ ' ' +  user.first_name +' '+user.last_name + ' ')

    return redirect('Allmodule',user.id)  
  return render(request,'company/register2.html')   

def staff_register(request):
  com=company.objects.all()

  return render(request, 'staff/staffreg.html',{'company':com})

def staff_registraction(request):
  if request.method == 'POST':
    fn=request.POST['fname']
    ln=request.POST['lname']
    email=request.POST['eid']
    un=request.POST['uname']
    pas=request.POST['pass']
    ph=request.POST['ph']
    code=request.POST['code']

    if company.objects.filter(Company_code=code).exists():
      com=company.objects.get(Company_code=code)
    else:
        messages.info(request, 'Sorry, Company code is Invalide')
        return redirect('staff_register')
    img=request.FILES.get('image')

    if staff_details.objects.filter(user_name=un).exists():
      messages.info(request, 'Sorry, Username already exists')
      return redirect('staff_register')
    elif staff_details.objects.filter(email=email).exists():
      messages.info(request, 'Sorry, Email already exists')
      return redirect('staff_register')
    else:
      
      staff=staff_details(first_name=fn,last_name=ln,email=email,user_name=un,password=pas,contact=ph,img=img,company=com)
      staff.save()
      return redirect('log_page')

  else:
    print(" error")
    return redirect('staff_register')
  





def companyaccept(request,id):
  data=staff_details.objects.filter(id=id).update(Action=1)
  return redirect('staff_request')

def companyreject(request,id):
  data=staff_details.objects.get(id=id)
  
  data.delete()
  return redirect('staff_request')

def client_request(request):
  data = company.objects.filter(superadmin_approval = 0,reg_action='self').order_by('-id')
  
  all = company.objects.filter(superadmin_approval = 1)
  return render(request,'admin/client_request.html',{'data': data,'all':all})

def client_request_overview(request,id): 
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  return render(request,'admin/client_request_overview.html',{'company':com,'allmodules':allmodules})

def client_details(request):
  data = company.objects.filter(superadmin_approval = 1,reg_action='self').order_by('-id')
  return render(request,'admin/client_details.html',{"data":data})
def client_details_overview(request,id): 
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  return render(request,'admin/client_details_overview.html',{'company':com,'allmodules':allmodules})



def payment_term(request):
  terms = payment_terms.objects.all()
  
  return render(request,'admin/payment_terms.html',{'terms':terms})
def add_payment_terms(request):
  if request.method == 'POST':
    num=int(request.POST['num'])
    select=request.POST['select']
    if select == 'Years':
      days=int(num)*365
      pt = payment_terms(payment_terms_number = num,payment_terms_value = select,days = days)
      pt.save()
      messages.info(request, 'Payment term is added')
      return redirect('payment_term')

    else:  
      days=int(num*30)
      pt = payment_terms(payment_terms_number = num,payment_terms_value = select,days = days)
      pt.save()
      messages.info(request, 'Payment term is added')
      return redirect('payment_term')


  return redirect('payment_term')




 







def admin_notification(request):
  data= modules_list.objects.filter(update_action=1,status='Pending')

  return render(request,'admin/admin_notification.html',{'data':data})

def module_updation_details(request,mid):
  data= modules_list.objects.get(id=mid)
  allmodules= modules_list.objects.get(company=data.company,status='Pending')
  old_modules = modules_list.objects.get(company=data.company,status='New')

  return render(request,'admin/module_updation_details.html',{'data':data,'allmodules':allmodules,'old_modules':old_modules})

def module_updation_ok(request,mid):
  
  old=modules_list.objects.get(company=mid,status='New')
  old.delete()

  data=modules_list.objects.get(company=mid,status='Pending')  
  data.status='New'
  data.save()
  data1=modules_list.objects.filter(company=mid).update(update_action=0)
  return redirect('admin_notification')






def distributor_reg(request):
  terms=payment_terms.objects.all()
  return render(request,'distributor/distributor_reg.html',{'terms':terms})
def distributor_reg_action(request):
  if request.method == 'POST':
    first_name = request.POST['fname']
    last_name = request.POST['lname']
    user_name = request.POST['uname']
    email_id = request.POST['eid']
    mobile = request.POST['ph']
    passw = request.POST['pass']
    c_passw = request.POST['cpass']
    pic = request.FILES.get('image')

    select=request.POST['select']
    terms=payment_terms.objects.get(id=select)
    # c.dateperiod=terms
    start_date=date.today()
    days=int(terms.days)

    
    end= date.today() + timedelta(days=days)
    End_date=end

    code=get_random_string(length=6)
    if Distributors_details.objects.filter(distributor_id = code).exists():
       code=get_random_string(length=6)
  
    if passw == c_passw:
      if User.objects.filter(username = user_name).exists():
        messages.info(request, 'Sorry, Username already exists')
        return redirect('distributor_reg')
      

      elif not User.objects.filter(email = email_id).exists():
    
        user_data = User.objects.create_user(first_name = first_name,
                        last_name = last_name,
                        username = user_name,
                        email = email_id,
                        password = passw)
        user_data.save()
        
        data = User.objects.get(id = user_data.id)
        distributor_data = Distributors_details(contact=mobile,distributor_id=code,img=pic,
                                                payment_term=terms,start_date=start_date,End_date=End_date,
                                                user = data)
        distributor_data.save()
        
        return redirect('log_page')
      else:
        messages.info(request, 'Sorry, Email already exists')
        return redirect('distributor_reg')
  return render(request,'distributor/distributor_reg.html')
 
def distributor_home(request):
  distributor =  Distributors_details.objects.get(user = request.user)

  return render(request,'distributor/distributor_home.html',{'distributor':distributor})



      
def clients(request):
  return render(request,'admin/clients.html')

def distributors(request):
  return render(request,'admin/distributors.html')  

def distributor_request(request):
  data = Distributors_details.objects.filter(Log_Action = 0).order_by('-id')
  return render(request,'admin/distributor_request.html',{'data':data})

def admin_distributor_accept(request,id):
  data=Distributors_details.objects.filter(id=id).update(Log_Action=1)
  return redirect('distributor_request')
def admin_distributor_reject(request,id):
  data=Distributors_details.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('distributor_request')

def distributor_request_overview(request,id):
  data=Distributors_details.objects.get(id=id)
  return render(request,'admin/distributor_request_overview.html',{'data':data})

def distributor_details(request):
  data = Distributors_details.objects.filter(Log_Action = 1).order_by('-id')
  return render(request,'admin/distributor_details.html',{'data':data})

def distributor_details_overview(request,id):
  data = Distributors_details.objects.get(id=id)
  return render(request,'admin/distributor_details_overview.html',{'data':data})

def dcompany_request(request):
  
  distributor =  Distributors_details.objects.get(user = request.user)
  data = company.objects.filter(Distributors = distributor,Distributor_approval = 0,reg_action='distributor').order_by('-id')
  return render(request,'distributor/dcompany_request.html',{'data':data,'distributor':distributor})

def dcompany_request_overview(request,id):
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/dcompany_request_overview.html',{'company':com,'allmodules':allmodules,'distributor':distributor})

def distributor_accept_company(request,id):
  data=company.objects.filter(id=id).update(Distributor_approval=1)
  
  return redirect('dcompany_request')
def distributor_reject_company(request,id):
  data=company.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('dcompany_request')

def dcompany_details(request):
  distributor =  Distributors_details.objects.get(user = request.user)
  data = company.objects.filter(Distributors = distributor,Distributor_approval = 1,reg_action='distributor').order_by('-id')
  
  return render(request,'distributor/dcompany_details.html',{'data':data,'distributor':distributor})

def dcompany_details_overview(request,id):
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/dcompany_details_overview.html',{'company':com,'allmodules':allmodules,'distributor':distributor})

def distributor_profile(request):
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/distributor_profile.html',{'distributor':distributor})

# ========================================   ASHIKH V U (START) ======================================================

@login_required(login_url='login')
def item_create(request):
  item_units = UnitModel.objects.filter(user=request.user.id)
  return render(request,'company/item_create.html',{'item_units':item_units})

@login_required(login_url='login')
def items_list(request,pk):
  try:
    get_company_id_using_user_id = company.objects.get(user=request.user.id)
    all_items = ItemModel.objects.filter(company=get_company_id_using_user_id.id)
    if pk == 0:
      first_item = all_items.filter().first()
    else:
      first_item = all_items.get(id=pk)
    transactions = TransactionModel.objects.filter(user=request.user.id,item=first_item.id).order_by('-trans_created_date')
    check_var = 0
    if all_items == None or all_items == '' or first_item == None or first_item == '' or transactions == None or transactions == '':
      return render(request,'company/items_create_first_item.html')
    return render(request,'company/items_list.html',{'all_items':all_items,
                                                      'first_item':first_item,
                                                      'transactions':transactions,})
  except:
    return render(request,'company/items_create_first_item.html')

@login_required(login_url='login')
def item_create_new(request):
  if request.method=='POST':
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_name = request.POST.get('item_name')
    item_hsn = request.POST.get('item_hsn')
    item_unit = request.POST.get('item_unit')
    item_taxable = request.POST.get('item_taxable')
    item_gst = request.POST.get('item_gst')
    item_igst = request.POST.get('item_igst')
    item_sale_price = request.POST.get('item_sale_price')
    item_purchase_price = request.POST.get('item_purchase_price')
    item_opening_stock = request.POST.get('item_opening_stock')
    item_current_stock = item_opening_stock
    if item_opening_stock == '' or None :
      item_opening_stock = 0
      item_current_stock = 0
    item_at_price = request.POST.get('item_at_price')
    if item_at_price == '' or None:
      item_at_price =0
    item_date = request.POST.get('item_date')
    item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
    if item_min_stock_maintain == ''  or None:
      item_min_stock_maintain = 0
    item_data = ItemModel(user=user,
                          company=company_user_data,
                          item_name=item_name,
                          item_hsn=item_hsn,
                          item_unit=item_unit,
                          item_taxable=item_taxable,
                          item_gst=item_gst,
                          item_igst=item_igst,
                          item_sale_price=item_sale_price,
                          item_purchase_price=item_purchase_price,
                          item_opening_stock=item_opening_stock,
                          item_current_stock=item_current_stock,
                          item_at_price=item_at_price,
                          item_date=item_date,
                          item_min_stock_maintain=item_min_stock_maintain)
    item_data.save()
    print(f'user : {user}\ncompany_user_data {company_user_data}')
    # print(f'item_name : {item_name}\nitem_hsn : {item_hsn}\nitem_unit : {item_unit}\nitem_taxable : {item_taxable}\n')
    # print(f'item_gst : {item_gst}\nitem_igst : {item_igst}\nitem_sale_price : {item_sale_price}\nitem_purchase_price : {item_purchase_price}\n')
    # print(f'item_opening_stock : {item_opening_stock}\nitem_at_price : {item_at_price}\nitem_date : {item_date}\nitem_min_stock_maintain : {item_min_stock_maintain}\n')
    print(f"----------\n\n\n")
    if request.POST.get('save_and_next'):
      return redirect('item_create')
    elif request.POST.get('save'):
      return redirect('items_list',pk=item_data.id)
  return redirect('item_create')


@login_required(login_url='login')
def item_delete(request,pk):
  get_company_id_using_user_id = company.objects.get(user=request.user.id)
  item_to_delete = ItemModel.objects.get(id=pk)
  item_to_delete.delete()
  return redirect('items_list',pk=0)


@login_required(login_url='login')
def item_view_or_edit(request,pk):
  item = ItemModel.objects.get(id=pk)
  item_units = UnitModel.objects.filter(user=request.user.id)
  return render(request,'company/item_view_or_edit.html',{'item':item,
                                                          'item_units':item_units,})

  
@login_required(login_url='login')
def item_unit_create(request):
  if request.method=='POST':
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_unit_name = request.POST.get('item_unit_name')
    unit_data = UnitModel(user=user,company=company_user_data,unit_name=item_unit_name)
    unit_data.save()
  return JsonResponse({'message':'asdasd'})

  
@login_required(login_url='login')
def item_update(request,pk):
  if request.method=='POST':
    item_data = ItemModel.objects.get(id=pk)
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_name = request.POST.get('item_name')
    item_hsn = request.POST.get('item_hsn')
    item_unit = request.POST.get('item_unit')
    item_taxable = request.POST.get('item_taxable')
    item_gst = request.POST.get('item_gst')
    item_igst = request.POST.get('item_igst')
    if item_taxable == 'Non Taxable':
      item_gst = 'GST0[0%]'
      item_igst = 'IGST0[0%]'
    item_sale_price = request.POST.get('item_sale_price')
    item_purchase_price = request.POST.get('item_purchase_price')
    item_opening_stock = request.POST.get('item_opening_stock')
    item_current_stock = item_opening_stock
    if item_opening_stock == '' :
      item_opening_stock = 0
      item_current_stock = 0
    else:
      if int(item_opening_stock) > item_data.item_opening_stock:
        item_data.item_current_stock += (int(item_opening_stock) - item_data.item_opening_stock)
      else:
        item_data.item_current_stock -= (int(item_opening_stock) - item_data.item_opening_stock)
    item_at_price = request.POST.get('item_at_price')
    if item_at_price == '':
      item_at_price =0
    item_date = request.POST.get('item_date')
    item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
    if item_min_stock_maintain == '':
      item_min_stock_maintain = 0

    item_data.user = user
    item_data.company_user_data = company_user_data
    item_data.item_name = item_name
    item_data.item_hsn = item_hsn
    item_data.item_unit = item_unit
    item_data.item_taxable = item_taxable
    item_data.item_gst = item_gst
    item_data.item_igst = item_igst
    item_data.item_sale_price = item_sale_price
    item_data.item_purchase_price = item_purchase_price
    item_data.item_opening_stock = item_opening_stock
    item_data.item_current_stock = int(item_current_stock)
    item_data.item_at_price = item_at_price
    item_data.item_date = item_date
    item_data.item_min_stock_maintain = item_min_stock_maintain

    item_data.save()
    print('\nupdated')
  # return redirect('item_view_or_edit',pk)
  return redirect('items_list',pk=item_data.id)

  
@login_required(login_url='login')
def item_search_filter(request):
  search_string = request.POST.get('searching_item')
  items_filtered = ItemModel.objects.filter(user=request.user.id)
  items_filtered = items_filtered.filter(Q(item_name__icontains=search_string))
  item_unit_name = request.POST.get('item_unit_name')
  return TemplateResponse(request,'company/item_search_filter.html',{'all_items':items_filtered})


@login_required(login_url='login')
def item_get_detail(request,pk):
  item = ItemModel.objects.get(id=pk)
  transactions = TransactionModel.objects.filter(user=request.user.id,item=item.id).order_by('-trans_created_date')
  return TemplateResponse(request,'company/item_get_detail.html',{"item":item,
                                                                  'transactions':transactions,})

  
@login_required(login_url='login')
def item_get_details_for_modal_target(request,pk):
  item = ItemModel.objects.get(id=pk)
  return TemplateResponse(request,'company/item_get_details_for_modal_target.html',{"item":item,})


@login_required(login_url='login')
def ajust_quantity(request,pk):
  if request.method=='POST':
    item = ItemModel.objects.get(id=pk)

    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    trans_type_check_checked = request.POST.get('trans_type')
    if trans_type_check_checked == 'on':
      trans_type = 'reduce stock'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'add stock'
      trans_qty = request.POST.get('added_qty')
    trans_user_name = user.first_name
    trans_date = request.POST.get('trans_date')

    trans_adjusted_qty= request.POST.get('adjusted_qty')
    trans_current_qty = request.POST.get('item_qty')
    print(f'the quantity : {trans_current_qty}')
    item.item_current_stock = trans_adjusted_qty
    item.save()
    transaction_data = TransactionModel(user=user,
                                        company=company_user_data,
                                        item=item,
                                        trans_type=trans_type,
                                        trans_user_name=trans_user_name,
                                        trans_date=trans_date,
                                        trans_qty=trans_qty,
                                        trans_current_qty=trans_current_qty,
                                        trans_adjusted_qty=trans_adjusted_qty,)
    transaction_data.save()
  return redirect('items_list',pk=item.id)


@login_required(login_url='login')
def transaction_delete(request,pk):
  transaction = TransactionModel.objects.get(id=pk)
  item = ItemModel.objects.get(id=transaction.item_id)
  print(transaction.trans_type)
  if transaction.trans_type=='add stock':
    print('add')
    item.item_current_stock = item.item_current_stock - transaction.trans_qty
    print(item.item_name)
    print(item.item_current_stock)
    print(item.item_current_stock)
    print(transaction.trans_qty)
    print(item.item_current_stock - transaction.trans_qty)
  else:
    print('reduce')
    item.item_current_stock = item.item_current_stock + transaction.trans_qty
  item.save()
  transaction.delete()
  return redirect('items_list',pk=item.id)

  
@login_required(login_url='login')
def item_transaction_view_or_edit(request,pk,tran):
  item = ItemModel.objects.get(id=pk)
  transaction = TransactionModel.objects.get(id=tran)
  print('enterd')
  return TemplateResponse(request,'company/item_transaction_view_or_edit.html',{"item":item,
                                                                                "transaction":transaction,})


@login_required(login_url='login')
def update_adjusted_transaction(request,pk,tran):
  item = ItemModel.objects.get(id=pk)
  transaction = TransactionModel.objects.get(id=tran)
  if request.method=='POST':
    item = ItemModel.objects.get(id=pk)

    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    trans_type_check_checked = request.POST.get('trans_type')
    if trans_type_check_checked == 'on':
      trans_type = 'reduce stock'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'add stock'
      trans_qty = request.POST.get('added_qty')
    trans_user_name = user.first_name
    trans_date = request.POST.get('trans_date')

    adjusted_qty= request.POST.get('adjusted_qty')
    trans_current_qty = request.POST.get('item_qty')
    if transaction.trans_type == 'reduce stock':
      if trans_type == 'reduce stock':
        print('reduce to reduce')
        item.item_current_stock = item.item_current_stock - (int(trans_qty)  - transaction.trans_qty)
      else:
        print('reduce to add')
        print(f'{trans_qty}-{transaction.trans_qty}={((int(trans_qty)  - transaction.trans_qty))}')
        item.item_current_stock = item.item_current_stock + transaction.trans_qty + int(trans_qty)
    else:
      if trans_type == 'reduce stock':
        print('add to red')
        item.item_current_stock = item.item_current_stock - (int(trans_qty)  + transaction.trans_qty)
      else:
        print('add to add')
        print(f'{trans_qty}-{transaction.trans_qty}={((int(trans_qty)  - transaction.trans_qty))}')
        item.item_current_stock = item.item_current_stock - transaction.trans_qty + int(trans_qty)
    # item.item_opening_stock = adjusted_qty
    item.save()
    transaction.trans_type =trans_type
    transaction.trans_date=trans_date
    transaction.trans_qty =trans_qty
    transaction.trans_current_qty=trans_current_qty
    transaction.save()
  return redirect('items_list',pk=item.id)
  
@login_required(login_url='login')
def item_delete_open_stk(request,pk):
  item = ItemModel.objects.get(id=pk)
  if item.item_opening_stock > item.item_current_stock:
    item.item_current_stock =item.item_opening_stock - item.item_current_stock
  else:
    item.item_current_stock =item.item_current_stock - item.item_opening_stock
  # item.item_current_stock =  item.item_opening_stock - item.item_current_stock
  item.item_opening_stock = 0
  # print(f'{item.item_current_stock }={item.item_opening_stock}-{item.item_current_stock}')
  item.save()
  return redirect('items_list',pk=item.id)
  
# ========================================   ASHIKH V U (END) ======================================================

#_________________Parties(new)_______________Antony Tom_________


def add_parties(request):
  return render(request, 'company/add_parties.html')










def edit_party(request,id):
  Company = company.objects.get(user=request.user)
  user_id = request.user.id
  getparty=party.objects.get(id=id)
  Party=party.objects.filter(user=request.user)
  return render(request, 'company/edit_party.html',{'Company':Company,'user_id':user_id,'Party':Party,'getparty':getparty})


def edit_saveparty(request, id):
    Party=party.objects.filter(user=request.user)
    user_id = request.user.id
    getparty = party.objects.get(id=id)
    Company = company.objects.get(user=request.user)

    if request.method == 'POST':
        getparty.party_name = request.POST.get('partyname')
        getparty.gst_no = request.POST.get('gstno')
        getparty.contact = request.POST['contact']
        getparty.gst_type = request.POST['gst']
        getparty.state = request.POST['state']
        getparty.address = request.POST['address']
        getparty.email = request.POST['email']
        getparty.openingbalance = request.POST['balance']
        getparty.payment = request.POST.get('paymentType')
        getparty.creditlimit = request.POST['creditlimit']
        getparty.current_date = request.POST['currentdate']
        getparty.additionalfield1 = request.POST['additionalfield1']
        getparty.additionalfield2 = request.POST['additionalfield2']
        getparty.additionalfield3 = request.POST['additionalfield3']

        getparty.save()

        return redirect('view_party', id=getparty.id)

    return render(request, 'edit_party.html', {'getparty': getparty, 'Party': Party, 'Company': Company,'user_id':user_id})


def deleteparty(request,id):
    Party=party.objects.get(id=id)
    Party.delete()
    return redirect('view_parties')

#End

@login_required(login_url='login')
def adminhome(request):
 
  
  
  return render(request, 'admin/adminhome.html')


#******************************************   ASHIKH V U (start) ****************************************************

from django.http import HttpResponse
import re

# account number validation
def validate_bank_account_number(acc_num):
  regex='^[0-9]{9,18}'
  if re.match(regex,acc_num):
    return True
  else:
    return False

# ifsc code validaion
def validate_ifsc(ifsc_code):
    regex = re.compile(r'^[A-Za-z]{4}\d{7}$')
    if regex.match(ifsc_code):
        return True
    else:
        return False

#@login_required(login_url='login')
def account_num_check(request):
  if request.method=='POST':
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    account_num_valid = validate_bank_account_number(account_num)
    if account_num_valid:
      if BankModel.objects.filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
        return HttpResponse('<small><span class="tr fs-2">Account Number already excist</span></small>')
      else:
        return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">Account Number is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def account_num_check_for_edit(request,pk):
  if request.method=='POST':
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    account_num_valid = validate_bank_account_number(account_num)
    if account_num_valid:
      if BankModel.objects.exclude(id=pk).filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
        return HttpResponse('<small><span class="tr fs-2">Account Number already excist</span></small>')
      else:
        return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">Account Number is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def bank_ifsc_check (request):
  if request.method=='POST':
    bank_ifsc = request.POST.get('ifsc')
    print(bank_ifsc)
    ifsc_valid = validate_ifsc(bank_ifsc)
    if ifsc_valid:
      return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">IFSC Code is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def bank_create(request):
  print('asdasd')
  try:
    staff_id = request.session['staff_id']
    print(staff_id)
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    # permission
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    # permission
    return render(request,'company/bank_create.html',{"allmodules":allmodules})
  except:
    user = User.objects.get(id=request.user.id)
    get_company_id_using_user_id = company.objects.get(user=user)
    # permission
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    # permission
    return render(request,'company/bank_create.html',{"allmodules":allmodules})


#@login_required(login_url='login')
def banks_list(request,pk):
  
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
 

  try:
    all_banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
    if pk == 0:
      first_bank = all_banks.first()
      print(all_banks)
      return redirect('banks_list',pk=first_bank.id)
    else:
      bank = all_banks.get(id=pk)
      transactions_all = BankTransactionModel.objects.filter(company=get_company_id_using_user_id.id)
      transactions = transactions_all.filter(Q(from_here=pk) | Q(to_here=pk))
      tr_history = BankTransactionHistory.objects.filter().order_by('date')
    if all_banks.exists():
      open_bal_last_edited = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED')).last()
      
      if bank.open_balance:
        total = bank.open_balance
      else:
        total = 0
      for i in transactions:
        if i.type == "Cash Withdraw":
          total = total - i.amount
        elif  i.type == 'Adjustment Reduce':
          total = total - i.amount
        elif i.from_here == bank:
          total = total - i.amount
        else:
          total = total + i.amount
        i.current_amount = total
      
      return render(request,'company/banks_list.html',{"allmodules":allmodules,
                                                      "all_banks":all_banks,
                                                      "bank":bank,
                                                      "transactions":transactions,
                                                      "tr_history":tr_history,
                                                      "open_bal_last_edited":open_bal_last_edited,
                                                      "staff":staff}) 
    else:
      return render(request,'company/bank_create_first_bank.html',{"allmodules":allmodules,'staff':staff}) 
  except:
    return render(request,'company/bank_create_first_bank.html',{"allmodules":allmodules,'staff':staff}) 
    

#@login_required(login_url='login')
def get_bank_to_bank(request):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_bank_to_bank.html',{'banks':banks})

#@login_required(login_url='login')
def get_bank_to_cash(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_bank_to_cash.html',{'banks':banks})

#@login_required(login_url='login')
def get_cash_to_bank(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_cash_to_bank.html',{'banks':banks})

#@login_required(login_url='login')
def get_adjust_bank_balance(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_adjust_bank_balance.html',{'banks':banks})

#@login_required(login_url='login')
def bank_create_new(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    print(get_company_id_using_user_id)
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    if BankModel.objects.exclude(company=get_company_id_using_user_id.id).filter(bank_name=bank_name,user=user.id,account_num=account_num).exists():
      parmission_var = 0
    else:
      parmission_var = 1
    if validate_bank_account_number(account_num):
      parmission_var1 = 1
    else:
      parmission_var1 = 0
    ifsc = request.POST.get('ifsc')
    if validate_ifsc(ifsc):
      parmission_var2 = 1
    else:
      parmission_var2 = 0
    branch_name = request.POST['branch_name']
    upi_id = request.POST.get('upi_id')
    as_of_date = request.POST['as_of_date']
    card_type = request.POST.get('card_type')
    open_balance = request.POST['open_balance']
    
    if open_balance == '' or open_balance == None:
      open_balance = 0
    if card_type == "CREDIT":
      open_balance = int(open_balance)*-1
      
    if parmission_var == 1:
      if parmission_var1 == 1:
        if parmission_var2 == 1:
          bank_data = BankModel(user=user,
                                company=get_company_id_using_user_id,
                                bank_name=bank_name,
                                account_num=account_num,
                                ifsc=ifsc,
                                branch_name=branch_name,
                                upi_id=upi_id,
                                as_of_date=as_of_date,
                                card_type=card_type,
                                open_balance=open_balance,
                                current_balance=open_balance,
                                created_by=user.first_name)
          bank_data.save()
          tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                              bank=bank_data,
                                              action="BANK CREATION : "+bank_data.bank_name.upper(),
                                              done_by_name=staff.first_name,
                                              done_by=staff)
          tr_history.save()
          tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                              bank=bank_data,
                                              action="BANK OPEN BALANCE CREATED",
                                              done_by_name=staff.first_name,
                                              done_by=staff)
          tr_history.save()
          if request.POST.get('save_and_next'):
            messages.success(request,'Bank created successfully')
            return redirect('bank_create')
          else:
            return redirect('banks_list',pk=bank_data.id)
        else:
          messages.error(request,'IFSC CODE is not valid')
          return redirect('bank_create')
      else:
        messages.error(request,'Account number is not valid')
        return redirect('bank_create')
    else:
      messages.error(request,'Account number already exist')
      return redirect('bank_create')
  return redirect('banks_list',pk=bank_data.id)

#@login_required(login_url='login')
def bank_delete(request,pk):
  bank = BankModel.objects.get(id=pk)
  bank.delete()
  return redirect('banks_list',pk=0)

#@login_required(login_url='login')
def bank_view_or_edit(request,pk):
  bank = BankModel.objects.get(id=pk)
  return render(request,'company/bank_view_or_edit.html',{"bank":bank})

#@login_required(login_url='login')
def bank_update(request,pk):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    
    bank_data = BankModel.objects.get(id=pk)

    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    if BankModel.objects.exclude(id=pk).filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
      parmission_var = 0
    else:
      parmission_var = 1
    if validate_bank_account_number(account_num):
      parmission_var1 = 1
    else:
      parmission_var1 = 0
    ifsc = request.POST.get('ifsc')
    if validate_ifsc(ifsc):
      parmission_var2 = 1
    else:
      parmission_var2 = 0
    branch_name = request.POST['branch_name']
    upi_id = request.POST.get('upi_id')
    as_of_date = request.POST['as_of_date']
    card_type = request.POST.get('card_type')
    open_balance = request.POST['open_balance']
    
    if open_balance == '' or open_balance == None:
      open_balance = 0
    if card_type == "CREDIT":
      open_balance = int(open_balance)*-1
    if parmission_var == 1:
      if parmission_var1 == 1:
        if parmission_var2 == 1:
          bank_data.user = user
          bank_data.company = get_company_id_using_user_id
          bank_data.bank_name = bank_name
          bank_data.account_num = account_num
          bank_data.ifsc = ifsc
          bank_data.branch_name = branch_name
          bank_data.upi_id = upi_id
          bank_data.as_of_date = as_of_date
          bank_data.card_type = card_type

          if int(bank_data.open_balance) < int(open_balance):
            bank_data.current_balance = int(bank_data.current_balance) + (int(open_balance) - int(bank_data.open_balance))
          elif int(bank_data.open_balance) == int(open_balance):
            bank_data.current_balance = int(open_balance)
          elif int(bank_data.open_balance) > int(open_balance):
            bank_data.current_balance = int(bank_data.current_balance)- (int(bank_data.open_balance) - int(open_balance))

          if bank_data.open_balance != open_balance:
            validity = True
          else:
            validity = False
          old_val = bank_data.open_balance

          bank_data.open_balance = open_balance
          bank_data.user = user
          bank_data.save()

          if validity == True:
            tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank=bank_data,
                                          action="BANK OPEN BALANCE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
            tr_history.save()
        else:
          messages.error(request,'IFSC CODE is not valid')
          return redirect('bank_create')
      else:
        messages.error(request,'Account number is not valid')
        return redirect('bank_create')
    else:
      messages.error(request,'Account number already exist')
      return redirect('bank_create')
  return redirect('banks_list',pk=bank_data.id)


#@login_required(login_url='login')
def bank_to_bank_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    select2 = request.POST.get('to_here')
    to_here = BankModel.objects.get(id=select2)
    type = "BANK TO BANK"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date')
    
    bank1 = BankModel.objects.get(id=from_here.id)
    bank1.current_balance -= int(amount)
    bank1.save()
    bank2 = BankModel.objects.get(id=to_here.id)
    bank2.current_balance += int(amount)
    bank2.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        to_here=to_here,
                                        type=type,
                                        date=date,
                                        name=name,
                                        amount=amount,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK TO BANK TRANSACTION CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def bank_to_cash_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    type = "Cash Withdraw"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank1 = BankModel.objects.get(id=from_here.id)
    bank1.current_balance -= int(amount)
    bank1.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def cash_to_bank_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select2 = request.POST.get('to_here')
    to_here = BankModel.objects.get(id=select2)
    type = "Cash Deposit"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank2 = BankModel.objects.get(id=to_here.id)
    bank2.current_balance += int(amount)
    bank2.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        to_here=to_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank2,
                                        bank_trans=transaction_data,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=to_here.id)


#@login_required(login_url='login')
def get_adjust_bank_balance_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    inc_red = request.POST.get('inc_red')
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank1 = BankModel.objects.get(id=from_here.id)
    if inc_red == 'Increase Balance':
      bank1.current_balance += int(amount) 
      type = "Adjustment Increase"
    else:
      bank1.current_balance -= int(amount)
      type = "Adjustment Reduce"
    bank1.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK BALANCE "+type.upper()+" CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def delete_bank_open_balance(request,pk):
  bank = BankModel.objects.get(id=pk)
  bank.current_balance = bank.current_balance - bank.open_balance
  bank.open_balance = 0
  bank.save()
  if 'banks_list' in request.META.get('HTTP_REFERER',None):
    return redirect('banks_list',pk=pk)
  else:
    return redirect('bank_transaction_statement',bank_id=pk)

#@login_required(login_url='login')
def delete_bank_transaction(request,pk,bank_id):
  print(pk,bank_id)
  try:
    pk = request.POST.get('pk')
    bank_id = request.POST.get('bank_id')
    print(pk,bank_id)
  except:
    pk=pk
    bank_id=bank_id

  try:
    trans = BankTransactionModel.objects.get(id=pk)
    if trans.type == 'BANK TO BANK':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      bank2.current_balance -= trans.amount
      bank2.save()
      trans.delete()
      print('enterd')
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Cash Withdraw' or trans.type == 'CASH WITHDRAW':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Cash Deposit' or trans.type == 'CASH DEPOSIT':
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      bank2.current_balance -= trans.amount
      bank2.save()
      trans.delete()
      print('entered')
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Adjustment Increase' or trans.type == 'ADJUSTMENT INCREASE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance -= trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Adjustment Reduce' or trans.type == 'ADJUSTMENT REDUCE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
  except:
    return redirect('banks_list',pk=bank_id)
  return redirect('banks_list',pk=0)

#@login_required(login_url='login')
def view_or_edit_bank_transaction(request,pk,bank_id):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  transaction = BankTransactionModel.objects.get(id=pk)
  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  bank = BankModel.objects.get(id=bank_id)
  if transaction.type == "BANK TO BANK" or transaction.type == 'Bank to bank':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_to_bank_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Cash Withdraw' or transaction.type == 'Cash withdraw' or transaction.type == 'CASH WITHDRAW':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_to_cash_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Cash Deposit' or transaction.type == 'Cash deposit' or transaction.type == 'CASH DEPOSIT':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/cash_to_bank_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Adjustment Increase' or transaction.type == 'Adjustment increase' or transaction.type == 'Adjustment Reduce' or transaction.type == 'Adjustment reduce' or transaction.type == 'ADJUSTMENT INCREASE' or transaction.type == 'ADJUSTMENT REDUCE':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_adjust_bank_balance_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})

#@login_required(login_url='login')
def update_bank_transaction(request,pk,bank_id):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    amount = request.POST.get('amount')
    date = request.POST.get('date')
    trans = BankTransactionModel.objects.get(id=pk)
    trans.date = date
    if trans.type == 'BANK TO BANK':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      if trans.amount > int(amount):
        bank2.current_balance -= (trans.amount-int(amount))
      else:
        bank2.current_balance += (int(amount)-trans.amount)
      bank2.save()
      old_amount = trans.amount
      if old_amount != amount:
        validity =True
      else:
        validity =False
      trans.amount = amount
      trans.save()
      if validity == True:
        tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK TO BANK TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
        tr_history.save()
        trans.last_action='UPDATED'
        trans.by = staff.first_name
        trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=bank_id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Cash Withdraw' or trans.type == 'CASH WITHDRAW':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK TO CASH TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Cash Deposit'  or trans.type == 'CASH DEPOSIT':
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      if trans.amount > int(amount):
        bank2.current_balance -= (trans.amount-int(amount))
      else:
        bank2.current_balance += (int(amount)-trans.amount)
      bank2.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="CASH TO BANK TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.to_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Adjustment Increase' or trans.type == 'ADJUSTMENT INCREASE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance -= (trans.amount-int(amount))
      else:
        bank1.current_balance += (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK BALANCE ADJUSTMENT INCREASE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Adjustment Reduce' or trans.type == 'ADJUSTMENT REDUCE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK BALANCE ADJUSTMENT REDUCE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    return redirect('banks_list',pk=0)
  return redirect('banks_list',pk=0)

from openpyxl import load_workbook
from django.utils import timezone

#@login_required(login_url='login')
def import_from_excel(request,pk):
    current_datetime = timezone.now()
    date =  current_datetime.date()

    try:
      if request.method == "POST" and 'excel_file' in request.FILES:
        
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        get_company_id_using_user_id = company.objects.get(id=staff.company.id)
        user = get_company_id_using_user_id.user
        allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            TYPE, FROM, TO,NAME,DATE,AMOUNT,ACTION,BY = row

            if TYPE != None:
              TYPE = TYPE.upper()
            
            if AMOUNT != None:
              AMOUNT = AMOUNT.replace(' ','')
              AMOUNT = AMOUNT.replace('₹','')
              AMOUNT = AMOUNT.replace('-','')
              AMOUNT = AMOUNT.replace('+','')
              AMOUNT = int(float(AMOUNT))

            print(f'{TYPE}  {FROM}  {TO}    {NAME}  {DATE}  {AMOUNT}')
            
            if TYPE == "BANK TO BANK" or TYPE == 'Bank to bank':
              from_here = BankModel.objects.get(id=int(FROM))
              to_here = BankModel.objects.get(id=int(TO))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  to_here=to_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              to_here.current_balance += AMOUNT
              to_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  bank_trans=transaction,
                                                  action="BANK TO BANK TRANSACTION CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
              
            elif TYPE == 'Open. Balance' or TYPE == 'OPEN. BALANCE':
              from_here = BankModel.objects.get(id=int(FROM))
              if from_here.open_balance > AMOUNT:
                from_here.current_balance += from_here.open_balance - AMOUNT
              else:
                from_here.current_balance -= from_here.open_balance - AMOUNT
              from_here.open_balance = AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  action="BANK OPEN BALANCE CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Withdraw' or TYPE == 'Cash withdraw' or TYPE == 'CASH WITHDRAW':
              from_here = BankModel.objects.get(id=int(FROM))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Deposit' or TYPE == 'Cash deposit' or TYPE == 'CASH DEPOSIT':
              to_here = BankModel.objects.get(id=int(TO))
              to_here.current_balance += AMOUNT
              to_here.save()

              transaction = BankTransactionModel(user = user,
                                                  company=get_company_id_using_user_id,
                                                  to_here=to_here,
                                                  type=TYPE,
                                                  amount=AMOUNT,
                                                  date=DATE,
                                                  last_action='CREATED',
                                                  by = staff.first_name,
                                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=to_here,
                                        bank_trans=transaction,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Increase' or TYPE == 'Adjustment increase' or TYPE == 'ADJUSTMENT INCREASE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance += AMOUNT
              from_here.save()
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Reduce' or TYPE == 'Adjustment reduce' or TYPE == 'ADJUSTMENT REDUCE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance -= AMOUNT
              from_here.save()
              transaction = BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
    except:
      messages.warning(request,"Table field is missing / you are importing the wrong File.")
    return redirect('banks_list',pk=pk)

#@login_required(login_url='login')
def import_statement_from_excel(request,pk):
    current_datetime = timezone.now()
    date =  current_datetime.date()

    try:
      if request.method == "POST" and 'excel_file' in request.FILES:
        
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        get_company_id_using_user_id = company.objects.get(id=staff.company.id)
        user = get_company_id_using_user_id.user
        allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            TYPE, FROM, TO,NAME,DATE,AMOUNT,BALANCE = row
            # TYPE, FROM, TO,NAME,DATE,AMOUNT,BALANCE,ACTION,BY = row

            if TYPE != None:
              TYPE = TYPE.upper()
            
            if AMOUNT != None:
              AMOUNT = AMOUNT.replace(' ','')
              AMOUNT = AMOUNT.replace('₹','')
              AMOUNT = AMOUNT.replace('-','')
              AMOUNT = AMOUNT.replace('+','')
              AMOUNT = int(float(AMOUNT))

            print(f'{TYPE}  {FROM}  {TO}    {NAME}  {DATE}  {AMOUNT}')
            
            if TYPE == "BANK TO BANK" or TYPE == 'Bank to bank':
              from_here = BankModel.objects.get(id=int(FROM))
              to_here = BankModel.objects.get(id=int(TO))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  to_here=to_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              to_here.current_balance += AMOUNT
              to_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  bank_trans=transaction,
                                                  action="BANK TO BANK TRANSACTION CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
              
            elif TYPE == 'Open. Balance' or TYPE == 'OPEN. BALANCE':
              from_here = BankModel.objects.get(id=int(FROM))
              if from_here.open_balance > AMOUNT:
                from_here.current_balance += from_here.open_balance - AMOUNT
              else:
                from_here.current_balance -= from_here.open_balance - AMOUNT
              from_here.open_balance = AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  action="BANK OPEN BALANCE CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Withdraw' or TYPE == 'Cash withdraw' or TYPE == 'CASH WITHDRAW':
              from_here = BankModel.objects.get(id=int(FROM))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Deposit' or TYPE == 'Cash deposit' or TYPE == 'CASH DEPOSIT':
              to_here = BankModel.objects.get(id=int(TO))
              to_here.current_balance += AMOUNT
              to_here.save()

              transaction = BankTransactionModel(user = user,
                                                  company=get_company_id_using_user_id,
                                                  to_here=to_here,
                                                  type=TYPE,
                                                  amount=AMOUNT,
                                                  date=DATE,
                                                  last_action='CREATED',
                                                  by = staff.first_name,
                                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=to_here,
                                        bank_trans=transaction,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Increase' or TYPE == 'Adjustment increase' or TYPE == 'ADJUSTMENT INCREASE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance += AMOUNT
              from_here.save()
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Reduce' or TYPE == 'Adjustment reduce' or TYPE == 'ADJUSTMENT REDUCE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance -= AMOUNT
              from_here.save()
              transaction = BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
    except:
      messages.warning(request,"Table field is missing / you are importing the wrong File.")
    return redirect('bank_transaction_statement',bank_id=pk) 

#@login_required(login_url='login')
def transaction_history(request,pk,bank_id):
    
    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    all_banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)

    # tr_history1 = BankTransactionHistory.objects.filter(action__contains='BANK CREATION',bank=bank_id)
    tr_history2 = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED'),bank=bank_id)    
    tr_history = BankTransactionHistory.objects.filter(bank_trans=pk)
    if pk != 0:
      # tr_historys = tr_history | tr_history1
      tr_historys = tr_history
    else:
      # tr_historys = tr_history1 | tr_history1 | tr_history2
      tr_historys =  tr_history | tr_history2
    # print(tr_history)
    
    return render(request,'company/bank_transaction_history.html',{"allmodules":allmodules,
                                                                   "all_banks":all_banks,
                                                                    "tr_historys":tr_historys,
                                                                    "bank_id":bank_id,
                                                                    "staff":staff})


#@login_required(login_url='login')
def bank_transaction_statement(request,bank_id):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
  bank = BankModel.objects.get(id=bank_id)

  transactions_all = BankTransactionModel.objects.filter(company=get_company_id_using_user_id.id)
  transactions = transactions_all.filter(Q(from_here=bank_id) | Q(to_here=bank_id))
  tr_history = BankTransactionHistory.objects.filter().order_by('date')

  open_bal_last_edited = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED')).last()
  
  if bank.open_balance:
    total = bank.open_balance
  else:
    total = 0
  for i in transactions:
    if i.type == "Cash Withdraw":
      total = total - i.amount
    elif  i.type == 'Adjustment Reduce':
      total = total - i.amount
    elif i.from_here == bank:
      total = total - i.amount
    else:
      total = total + i.amount
    i.current_amount = total

  return render(request,'company/bank_transaction_statement.html',{"allmodules":allmodules,
                                                  "bank":bank,
                                                  "transactions":transactions,
                                                  "tr_history":tr_history,
                                                  "open_bal_last_edited":open_bal_last_edited,
                                                  "staff":staff})

#******************************************   ASHIKH V U (end) ****************************************************


def view_purchasebill(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allmodules= modules_list.objects.get(company=cmp,status='New')
  pbill = PurchaseBill.objects.filter(company=cmp)

  if not pbill:
    context = {'staff':staff, 'allmodules':allmodules}
    return render(request,'company/purchasebillempty.html',context)
  
  context = {'staff':staff,'allmodules':allmodules,'pbill':pbill}
  return render(request,'company/purchasebilllist.html',context)


def add_purchasebill(request):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  last_bill = PurchaseBill.objects.filter(company=cmp).last()

  if last_bill:
    bill_no = last_bill.tot_bill_no + 1 
  else:
    bill_no = 1

  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company)

  context = {'staff':staff, 'allmodules':allmodules, 'cust':cust, 'cmp':cmp,'bill_no':bill_no, 'tod':tod, 'item':item, 'item_units':item_units,'bank':bank}
  return render(request,'company/purchasebilladd.html',context)


def create_purchasebill(request):
  if request.method == 'POST': 
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)    
    part = party.objects.get(id=request.POST.get('customername'))
    print('part', part)
    pbill = PurchaseBill(party=part, 
                          billno=request.POST.get('bill_no'),
                          billdate=request.POST.get('billdate'),
                          supplyplace =request.POST.get('placosupply'),
                          pay_method=request.POST.get("method"),
                          cheque_no=request.POST.get("cheque_id"),
                          upi_no=request.POST.get("upi_id"),
                          advance = request.POST.get("advance"),
                          balance = request.POST.get("balance"),
                          subtotal=float(request.POST.get('subtotal')),
                          igst = request.POST.get('igst'),
                          cgst = request.POST.get('cgst'),
                          sgst = request.POST.get('sgst'),
                          adjust = request.POST.get("adj"),
                          taxamount = request.POST.get("taxamount"),
                          grandtotal=request.POST.get('grandtotal'),
                          company=cmp,staff=staff)
    pbill.save()
        
    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    if request.POST.get('placosupply') =='State':
      tax =  tuple(request.POST.getlist("tax1[]"))
    else:
      tax =  tuple(request.POST.getlist("tax2[]"))
    total =  tuple(request.POST.getlist("total[]"))
    billno = PurchaseBill.objects.get(billno =pbill.billno,company=cmp)

    if len(product)==len(qty)==len(tax)==len(discount)==len(total):
        mapped=zip(product,qty,tax,discount,total)
        mapped=list(mapped)
        for ele in mapped:
          itm = ItemModel.objects.get(id=ele[0])
          PurchaseBillItem.objects.create(product = itm,qty=ele[1], tax=ele[2],discount=ele[3],total=ele[4],purchasebill=billno,company=cmp)

    PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=F('tot_bill_no') + 1)
    
    pbill.tot_bill_no = pbill.billno
    pbill.save()

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,company=cmp,staff=staff,action='Created')

    if 'Next' in request.POST:
      return redirect('add_purchasebill')
    
    if "Save" in request.POST:
      return redirect('view_purchasebill')
    
  else:
    return render(request,'staff/purchasebilladd.html')


def edit_purchasebill(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  billprd = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)

  if pbill.pay_method != 'Cash' and pbill.pay_method != 'Cheque' and pbill.pay_method != 'UPI':
    bankno = BankModel.objects.get(id= pbill.pay_method,company=cmp,user=cmp.user)
  else:
    bankno = 0

  bdate = pbill.billdate.strftime("%Y-%m-%d")
  context = {'staff':staff, 'allmodules':allmodules, 'pbill':pbill, 'billprd':billprd,'tod':tod,
             'cust':cust, 'item':item, 'item_units':item_units, 'bdate':bdate,'bank':bank, 'bankno':bankno}
  return render(request,'company/purchasebilledit.html',context)


def update_purchasebill(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    part = party.objects.get(id=request.POST.get('customername'))
    pbill = PurchaseBill.objects.get(id=id,company=cmp)
    pbill.party = part
    pbill.billdate = request.POST.get('billdate')
    pbill.supplyplace  = request.POST.get('placosupply')
    pbill.subtotal =float(request.POST.get('subtotal'))
    pbill.grandtotal = request.POST.get('grandtotal')
    pbill.igst = request.POST.get('igst')
    pbill.cgst = request.POST.get('cgst')
    pbill.sgst = request.POST.get('sgst')
    pbill.taxamount = request.POST.get("taxamount")
    pbill.adjust = request.POST.get("adj")
    pbill.pay_method = request.POST.get("method")
    pbill.cheque_no = request.POST.get("cheque_id")
    pbill.upi_no = request.POST.get("upi_id")
    pbill.advance = request.POST.get("advance")
    pbill.balance = request.POST.get("balance")

    pbill.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    if request.POST.get('placosupply') == 'State':
      tax =tuple( request.POST.getlist("tax1[]"))
    else:
      tax = tuple(request.POST.getlist("tax2[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
    if len(total)==len(discount)==len(qty)==len(tax):
      mapped=zip(product,qty,tax,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        PurchaseBillItem.objects.create(product =itm,qty=ele[1], tax=ele[2],discount=ele[3],total=ele[4],purchasebill=pbill,company=cmp)

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,company=cmp,staff=staff,action='Updated')
    return redirect('view_purchasebill')

  return redirect('view_purchasebill')


def details_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  allmodules = modules_list.objects.get(company=staff.company,status='New')
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  pitm = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)
  dis = 0
  for itm in pitm:
    dis += int(itm.discount)
  itm_len = len(pitm)

  context={'staff':staff,'allmodules':allmodules,'pbill':pbill,'pitm':pitm,'itm_len':itm_len,'dis':dis}
  return render(request,'company/purchasebilldetails.html',context)


def history_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  hst= PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp)

  context = {'staff':staff,'allmodules':allmodules,'hst':hst,'pbill':pbill}
  return render(request,'company/purchasebillhistory.html',context)


def delete_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pbill = PurchaseBill.objects.get(id=id)
  PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
  pbill.delete()
  return redirect('view_purchasebill')


def import_purchase_bill(request):
  if request.method == 'POST' and request.FILES['billfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(PurchaseBill.objects.filter(company=cmp).last().tot_bill_no) + 1

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=billsheet[0],email=billsheet[1],company=cmp)
      PurchaseBill.objects.create(party=part,billno=totval,
                                  billdate=billsheet[2],
                                  supplyplace =billsheet[3],
                                  tot_bill_no = totval,
                                  company=cmp,staff=staff)
      
      pbill = PurchaseBill.objects.last()
      if billsheet[4] == 'Cheque':
        pbill.pay_method = 'Cheque'
        pbill.cheque_no = billsheet[5]
      elif billsheet[4] == 'UPI':
        pbill.pay_method = 'UPI'
        pbill.upi_no = billsheet[5]
      else:
        if billsheet[4] != 'Cash':
          bank = BankModel.objects.get(bank_name=billsheet[4],company=cmp)
          pbill.pay_method = bank
        else:
          pbill.pay_method = 'Cash'
      pbill.save()

      PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=totval)
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=prdsheet[2])
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[5])
          PurchaseBillItem.objects.create(purchasebill=pbill,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                tax=prdsheet[4],
                                discount=prdsheet[5],
                                total=total)

          temp = prdsheet[4].split('[')
          if billsheet[3] =='State':
            tax=int(temp[0][3:])
          else:
            tax=int(temp[0][4:])

          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount
                
      if billsheet[3]=='State':
        gst = round((taxamount/2),2)
        pbill.sgst=gst
        pbill.cgst=gst
        pbill.igst=0

      else:
        gst=round(taxamount,2)
        pbill.igst=gst
        pbill.cgst=0
        pbill.sgst=0

      gtotal = subtotal + taxamount + float(billsheet[6])
      balance = gtotal- float(billsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pbill.subtotal=round(subtotal,2)
      pbill.taxamount=round(taxamount,2)
      pbill.adjust=round(billsheet[6],2)
      pbill.grandtotal=gtotal
      pbill.advance=round(billsheet[7],2)
      pbill.balance=balance
      pbill.save()

      PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,staff=pbill.staff,company=pbill.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})


def billhistory(request):
  pid = request.POST['id']
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pbill = PurchaseBill.objects.get(billno=pid,company=cmp)
  hst = PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp).last()
  name = hst.staff.first_name + ' ' + hst.staff.last_name 
  action = hst.action
  return JsonResponse({'name':name,'action':action,'pid':pid})


def bankdata(request):
  bid = request.POST['id']
  bank = BankModel.objects.get(id=bid) 
  bank_no = bank.account_num
  bank_name = bank.bank_name
  return JsonResponse({'bank_no':bank_no,'bank_name':bank_name})


def savecustomer(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  party_name = request.POST['name']
  email = request.POST['email']
  contact = request.POST['mobile']
  state = request.POST['splystate']
  address = request.POST['baddress']
  gst_type = request.POST['gsttype']
  gst_no = request.POST['gstin']
  current_date = request.POST['partydate']
  openingbalance = request.POST.get('openbalance')
  payment = request.POST.get('paytype')
  creditlimit = request.POST.get('credit_limit')
  End_date = request.POST.get('enddate', None)
  additionalfield1 = request.POST['add1']
  additionalfield2 = request.POST['add2']
  additionalfield3 = request.POST['add3']

  part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,
                payment=payment,creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,
                additionalfield3=additionalfield3,company=cmp,user=cmp.user)
  part.save() 
  return JsonResponse({'success': True})


def cust_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part = party.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  party_list = []
  for p in part:
    id_list.append(p.id)
    party_list.append(p.party_name)

  return JsonResponse({'id_list':id_list, 'party_list':party_list })


def saveitem(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  name = request.POST['name']
  unit = request.POST['unit']
  hsn = request.POST['hsn']
  taxref = request.POST['taxref']
  sell_price = request.POST['sell_price']
  cost_price = request.POST['cost_price']
  intra_st = request.POST['intra_st']
  inter_st = request.POST['inter_st']

  if taxref != 'Taxable':
    intra_st = 'GST0[0%]'
    inter_st = 'IGST0[0%]'

  itmdate = request.POST.get('itmdate')
  stock = request.POST.get('stock')
  itmprice = request.POST.get('itmprice')
  minstock = request.POST.get('minstock')

  if not hsn:
    hsn = None

  itm = ItemModel(item_name=name, item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
  itm.save() 
  return JsonResponse({'success': True})


def item_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  options = {}
  option_objects = ItemModel.objects.filter(company=cmp,user=cmp.user)
  for option in option_objects:
      options[option.id] = [option.item_name]
  return JsonResponse(options)


def custdata(request):
  cid = request.POST['id']
  part = party.objects.get(id=cid)
  phno = part.contact
  address = part.address
  pay = part.payment
  bal = part.openingbalance
  return JsonResponse({'phno':phno, 'address':address, 'pay':pay, 'bal':bal})


def itemdetails(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})

def add_purchaseorder(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pbill = PurchaseBill.objects.filter(company=staff.company)
  context = {'staff':staff, 'allmodules':allmodules}
  return render(request,'company/purchasebillempty.html',context)
  
  
def view_purchaseorder(request):
   sid = request.session.get('staff_id')
   staff =  staff_details.objects.get(id=sid)
   allmodules= modules_list.objects.get(company=staff.company,status='New')
   context = {'staff':staff, 'allmodules':allmodules}
   return render(request,'company/purchasebillempty.html',context)

# ===========  estimate & delivery challan ===========shemeem==================   
   
def delivery_challan(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    all_challan = DeliveryChallan.objects.filter(company = com)
    challan = []
    for dc in all_challan:
      history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
      dict = {'challan':dc,'history':history}
      challan.append(dict)
    context = {
      'staff':staff, 'company':com,'allmodules':allmodules, 'challan':challan,
    }
    return render(request, 'staff/delivery_challan.html',context)
    

def create_estimate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      allmodules= modules_list.objects.get(company=com.id,status='New')
      parties = party.objects.filter(company = com)
      items = ItemModel.objects.filter(company = com)
      item_units = UnitModel.objects.filter(company=com)

      # Fetching last bill and assigning upcoming bill no as current + 1
      # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
      latest_bill = Estimate.objects.filter(company = com).order_by('-id').first()

      if latest_bill:
          last_number = int(latest_bill.ref_no)
          new_number = last_number + 1
      else:
          new_number = 1

      if DeletedEstimate.objects.filter(company = com).exists():
          deleted = DeletedEstimate.objects.get(company = com)
          
          if deleted:
              while int(deleted.ref_no) >= new_number:
                  new_number+=1

      
      context = {
        'staff':staff, 'company':com,'allmodules':allmodules, 'parties':parties, 'ref_no':new_number,'items':items,'item_units':item_units,
      }
      return render(request, 'staff/create_estimate.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def getPartyDetails(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)  
    party_id = request.POST.get('id')
    party_details = party.objects.get(id = party_id)

    list = []
    dict = {
      'contact': party_details.contact,
      'address':party_details.address,
      'state': party_details.state,
      'balance':party_details.openingbalance,
      'payment':party_details.payment,
    }
    list.append(dict)
    return JsonResponse(json.dumps(list), content_type="application/json", safe=False)
    

def getItemData(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)  
    try:
        id = request.GET.get('id')

        # item = ItemModel.objects.get(item_name = id, company = com)
        item = ItemModel.objects.filter(item_name = id, company = com).first()
        hsn = item.item_hsn
        pur_rate = item.item_purchase_price
        sale_rate = item.item_sale_price
        tax = True if item.item_taxable == "Taxable" else False
        gst = item.item_gst
        igst = item.item_igst

        return JsonResponse({"status":True,'id':item.id,'hsn':hsn,'pur_rate':pur_rate,'sale_rate':sale_rate, 'tax':tax, 'gst':gst, 'igst':igst})
    except Exception as e:
        print(e)
        return JsonResponse({"status":False})
  

def createNewEstimate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
        if request.method == 'POST':
            estimate = Estimate(
              staff = staff,
              company = com,
              date = request.POST['date'],
              ref_no = request.POST['ref_no'],
              party_name = party.objects.get(id = request.POST['party_name']).party_name,
              contact = request.POST['contact'],
              billing_address = request.POST['address'],
              state_of_supply = 'State' if request.POST['state_supply'] == 'state' else 'Other State',
              description = request.POST['description'],
              subtotal = request.POST['subtotal'],
              cgst = request.POST['cgst_tax'],
              sgst = request.POST['sgst_tax'],
              igst = request.POST['igst_tax'],
              tax_amount = request.POST['tax_amount'],
              adjustment = request.POST['adjustment'],
              total_amount = request.POST['grand_total'],
              balance = 0,
              status = 'Open',
              is_converted = False
            )
            estimate.save()
            
            ids = request.POST.getlist('estItems[]')
            item = request.POST.getlist("item[]")
            hsn  = request.POST.getlist("hsn[]")
            qty = request.POST.getlist("qty[]")
            price = request.POST.getlist("price[]")
            tax = request.POST.getlist("taxgst[]") if request.POST['state_supply'] == 'state' else request.POST.getlist("taxigst[]")
            discount = request.POST.getlist("discount[]")
            total = request.POST.getlist("total[]")

            est_id = Estimate.objects.get( id = estimate.id)

            if len(ids)==len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total) and ids and item and hsn and qty and price and tax and discount and total:
                mapped = zip(ids,item,hsn,qty,price,tax,discount,total)
                mapped = list(mapped)
                for ele in mapped:
                  estItems = Estimate_items.objects.create(staff = staff, eid = est_id, company = com, item = ItemModel.objects.get(company = com, id = ele[0]),name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],discount = ele[6],total=ele[7])
            
            # Transaction history

            history = EstimateTransactionHistory(
              staff = staff,
              estimate = estimate,
              company = com,
              action = "Create"
            )
            history.save()

            if 'save_and_next' in request.POST:
                return redirect(create_estimate)
            return redirect(estimate_quotation)
    except Exception as e:
        print(e)
        return redirect(create_estimate)
  return redirect('/')


def getPartyList(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    options = {}
    option_objects = party.objects.filter(company = com)
    for option in option_objects:
        options[option.id] = [option.id , option.party_name]

    return JsonResponse(options)


def getItemList(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    options = {}
    option_objects = ItemModel.objects.filter(company = com)
    for option in option_objects:
        options[option.id] = [option.item_name]

    return JsonResponse(options)
  

def estimateFilterWithDate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      date = request.GET['date_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, date = date)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)      
      
      if not all_estimates:
        messages.warning(request, f'No Estimates found on {date}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
      }
      return render(request, 'staff/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    

def estimateFilterWithRef(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      ref = request.GET['ref_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, ref_no = ref)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Ref No. {ref}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
      }
      return render(request, 'staff/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def estimateFilterWithBal(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      bal = request.GET['bal_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, balance = bal)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Balance amount {bal}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
      }
      return render(request, 'staff/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def estimateFilterWithName(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      name = request.GET['name_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, party_name = name)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)      

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Party Name {name}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
      }
      return render(request, 'staff/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def estimateFilterWithTotal(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      tot = request.GET['total_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, total_amount = tot)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Total Amount {tot}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)

      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
      }
      return render(request, 'staff/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    
  
def estimateFilterWithStat(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      stat = request.GET['status']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, status = stat)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Status {stat}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
      }
      return render(request, 'staff/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
   


def estimateInBetween(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      fromDate = request.GET['from_date']
      toDate = request.GET['to_date']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com).filter(date__gte = fromDate, date__lte = toDate)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)
      
      if not all_estimates:
        messages.warning(request, f'No Estimates found in between {fromDate} to {toDate}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)      
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
      }
      return render(request, 'staff/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def deleteEstimate(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      est = Estimate.objects.get(company = com, id = id)

      # Storing ref number to deleted table
      # if entry exists and lesser than the current, update and save => Only one entry per company

      if DeletedEstimate.objects.filter(company = com).exists():
          deleted = DeletedEstimate.objects.get(company = com)
          if deleted:
              if int(est.ref_no) > int(deleted.ref_no):
                  deleted.ref_no = est.ref_no
                  deleted.save()
          
      else:
          deleted = DeletedEstimate(company = com, staff = staff, ref_no = est.ref_no)
          deleted.save()
      
      Estimate_items.objects.filter(company = com , eid = est).delete()
      est.delete()
      messages.success(request, 'Estimate deleted successfully.!')
      return redirect(estimate_quotation)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
  return redirect('/')


def editEstimate(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      est = Estimate.objects.get(company = com , id = id)
      est_items = Estimate_items.objects.filter(company = com , eid = est)
      allmodules= modules_list.objects.get(company=com.id,status='New')
      parties = party.objects.filter(company = com)
      items = ItemModel.objects.filter(company = com)
      item_units = UnitModel.objects.filter(company=com)
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'parties':parties,'items':items,'item_units':item_units, 'estimate':est, 'estItems':est_items,
      }
      return render(request, 'staff/edit_estimate.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    

def updateEstimate(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      estimate = Estimate.objects.get(company = com, id = id)
      if request.method == 'POST':
        estimate.date = request.POST['date']
        estimate.ref_no = request.POST['ref_no']
        estimate.party_name = party.objects.get(id = request.POST['party_name']).party_name
        estimate.contact = request.POST['contact']
        estimate.billing_address = request.POST['address']
        estimate.state_of_supply = 'State' if request.POST['state_supply'] == 'state' else 'Other State'
        estimate.description = request.POST['description']
        estimate.subtotal = request.POST['subtotal']
        estimate.cgst = request.POST['cgst_tax']
        estimate.sgst = request.POST['sgst_tax']
        estimate.igst = request.POST['igst_tax']
        estimate.tax_amount = request.POST['tax_amount']
        estimate.adjustment = request.POST['adjustment']
        estimate.total_amount = request.POST['grand_total']
        estimate.balance = 0
        estimate.status = 'Open'
        estimate.is_converted = False

        estimate.save()

        ids = request.POST.getlist('estItems[]')
        item = request.POST.getlist("item[]")
        hsn  = request.POST.getlist("hsn[]")
        qty = request.POST.getlist("qty[]")
        price = request.POST.getlist("price[]")
        tax = request.POST.getlist("taxgst[]") if request.POST['state_supply'] == 'state' else request.POST.getlist("taxigst[]")
        discount = request.POST.getlist("discount[]")
        total = request.POST.getlist("total[]")
        est_item_ids = request.POST.getlist("id[]")
        
        item_ids = [int(id) for id in est_item_ids]

        
        est_item = Estimate_items.objects.filter(eid = estimate)
        object_ids = [obj.id for obj in est_item]

        ids_to_delete = [obj_id for obj_id in object_ids if obj_id not in item_ids]

        Estimate_items.objects.filter(id__in=ids_to_delete).delete()
        
        count = Estimate_items.objects.filter(eid = estimate, company = com).count()
        if len(ids)==len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total):
            try:
                mapped=zip(ids,item,hsn,qty,price,tax,total,discount,item_ids)
                mapped=list(mapped)
                
                for ele in mapped:
                    if int(len(item))>int(count):
                        if ele[8] == 0:
                            itemAdd= Estimate_items.objects.create(name = ele[1], hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7] ,staff = staff ,eid = estimate ,company = com, item = ItemModel.objects.get(company = com, id = ele[0]))
                        else:
                            itemAdd = Estimate_items.objects.filter( id = ele[8],company = com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
                    else:
                        itemAdd = Estimate_items.objects.filter( id = ele[8],company=com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
            except Exception as e:
                    print(e)
                    mapped=zip(ids,item,hsn,qty,price,tax,total,discount,item_ids)
                    mapped=list(mapped)
                    
                    for ele in mapped:
                        created =Estimate_items.objects.filter(id=ele[8] ,company=com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
        
        history = EstimateTransactionHistory(
          staff = staff,
          estimate = estimate,
          company = com,
          action = "Edit"
        )
        history.save()

        return redirect(viewEstimate,id)
    except Exception as e:
      print(e)
      return redirect(editEstimate, id)
    

def estimateTransactionHistory(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    try:
      est = Estimate.objects.get(company = com, id = id)
      history = EstimateTransactionHistory.objects.filter(company = com, estimate = est)
      context = {
        'staff':staff, 'company':com,'allmodules':allmodules,'history':history,
      }
      return render(request, 'staff/estimate_transaction_history.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    

# DELIVERY CHALLAN

def createDeliveryChallan(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      allmodules= modules_list.objects.get(company=com.id,status='New')
      parties = party.objects.filter(company = com)
      items = ItemModel.objects.filter(company = com)
      item_units = UnitModel.objects.filter(company=com)

      # Fetching last bill and assigning upcoming bill no as current + 1
      # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
      latest_bill = DeliveryChallan.objects.filter(company = com).order_by('-id').first()

      if latest_bill:
          last_number = int(latest_bill.challan_no)
          new_number = last_number + 1
      else:
          new_number = 1

      if DeletedDeliveryChallan.objects.filter(company = com).exists():
          deleted = DeletedDeliveryChallan.objects.get(company = com)
          
          if deleted:
              while int(deleted.challan_no) >= new_number:
                  new_number+=1

      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'parties':parties, 'challan_no':new_number,'items':items,'item_units':item_units,
      }
      return render(request, 'staff/create_delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def createNewDeliveryChallan(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
        if request.method == 'POST':
            challan = DeliveryChallan(
              company = com,
              staff = staff,
              date = request.POST['date'],
              due_date = request.POST['due_date'],
              challan_no = request.POST['challan_no'],
              party_name = party.objects.get(id = request.POST['party_name']).party_name,
              contact = request.POST['contact'],
              billing_address = request.POST['address'],
              state_of_supply = 'State' if request.POST['state_supply'] == 'state' else 'Other State',
              description = request.POST['description'],
              subtotal = request.POST['subtotal'],
              cgst = request.POST['cgst_tax'],
              sgst = request.POST['sgst_tax'],
              igst = request.POST['igst_tax'],
              tax_amount = request.POST['tax_amount'],
              adjustment = request.POST['adjustment'],
              total_amount = request.POST['grand_total'],
              balance = 0,
              status = 'Open',
              is_converted = False
            )
            challan.save()
            
            ids = request.POST.getlist('dcItems[]')
            item = request.POST.getlist("item[]")
            hsn  = request.POST.getlist("hsn[]")
            qty = request.POST.getlist("qty[]")
            price = request.POST.getlist("price[]")
            tax = request.POST.getlist("taxgst[]") if request.POST['state_supply'] == 'state' else request.POST.getlist("taxigst[]")
            discount = request.POST.getlist("discount[]")
            total = request.POST.getlist("total[]")

            chl_id = DeliveryChallan.objects.get( id = challan.id)

            if len(ids)==len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total) and ids and item and hsn and qty and price and tax and discount and total:
                mapped = zip(ids,item,hsn,qty,price,tax,discount,total)
                mapped = list(mapped)
                for ele in mapped:
                  dcItems = DeliveryChallanItems.objects.create(staff = staff,cid = chl_id, company = com, item = ItemModel.objects.get(company = com, id = ele[0]),name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],discount = ele[6],total=ele[7])
            
            history = DeliveryChallanTransactionHistory(
              staff = staff,
              challan = challan,
              company = com,
              action = "Create"
            )
            history.save()

            if 'save_and_next' in request.POST:
                return redirect(createDeliveryChallan)
            return redirect(delivery_challan)
    except Exception as e:
        print(e)
        return redirect(createDeliveryChallan)
  return redirect('/')


def challanInBetween(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      fromDate = request.GET['from_date']
      toDate = request.GET['to_date']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com).filter(date__gte = fromDate, date__lte = toDate)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)
      if not all_challan:
        messages.warning(request, f'No Challans found in between {fromDate} to {toDate}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,
      }
      return render(request, 'staff/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def challanFilterWithDate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      date = request.GET['date_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, date = date)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)
      if not all_challan:
        messages.warning(request, f'No Challans found on {date}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,
      }
      return render(request, 'staff/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
    

def challanFilterWithDueDate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      date = request.GET['due_date_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, due_date = date)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Due Date {date}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,
      }
      return render(request, 'staff/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
    

def challanFilterWithChallanNo(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      chl = request.GET['challan_no_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, challan_no = chl)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Challan No. {chl}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,
      }
      return render(request, 'staff/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def challanFilterWithBal(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      bal = request.GET['bal_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, balance = bal)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Balance amount {bal}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,
      }
      return render(request, 'staff/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def challanFilterWithName(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      name = request.GET['name_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, party_name = name)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Party Name {name}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,
      }
      return render(request, 'staff/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def challanFilterWithTotal(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      tot = request.GET['total_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, total_amount = tot)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Total Amount {tot}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)

      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,
      }
      return render(request, 'staff/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
    
  
def challanFilterWithStat(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      stat = request.GET['status']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, status = stat)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)
        
      if not all_challan:
        messages.warning(request, f'No Challans found with Status {stat}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,
      }
      return render(request, 'staff/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def deleteChallan(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      challan = DeliveryChallan.objects.get(company = com, id = id)

      # Storing ref number to deleted table
      # if entry exists and lesser than the current, update and save => Only one entry per company

      if DeletedDeliveryChallan.objects.filter(company = com).exists():
          deleted = DeletedDeliveryChallan.objects.get(company = com)
          if deleted:
              if int(challan.challan_no) > int(deleted.challan_no):
                  deleted.challan_no = challan.challan_no
                  deleted.save()
          
      else:
          deleted = DeletedDeliveryChallan(company = com, staff = staff, challan_no = challan.challan_no)
          deleted.save()
      
      DeliveryChallanItems.objects.filter(company = com , cid = challan).delete()
      challan.delete()
      messages.success(request, 'Challan deleted successfully.!')
      return redirect(delivery_challan)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
  return redirect('/')


def editChallan(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      dc = DeliveryChallan.objects.get(company = com , id = id)
      dc_items = DeliveryChallanItems.objects.filter(company = com , cid = dc)
      allmodules= modules_list.objects.get(company=com.id,status='New')
      parties = party.objects.filter(company = com)
      items = ItemModel.objects.filter(company = com)
      item_units = UnitModel.objects.filter(company=com)
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'parties':parties,'items':items,'item_units':item_units, 'challan':dc, 'dcItems':dc_items,
      }
      return render(request, 'staff/edit_delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)



def updateChallan(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      challan = DeliveryChallan.objects.get(company = com, id = id)
      if request.method == 'POST':
        challan.date = request.POST['date']
        challan.due_date = request.POST['due_date']
        challan.challan_no = request.POST['challan_no']
        challan.party_name = party.objects.get(id = request.POST['party_name']).party_name
        challan.contact = request.POST['contact']
        challan.billing_address = request.POST['address']
        challan.state_of_supply = 'State' if request.POST['state_supply'] == 'state' else 'Other State'
        challan.description = request.POST['description']
        challan.subtotal = request.POST['subtotal']
        challan.cgst = request.POST['cgst_tax']
        challan.sgst = request.POST['sgst_tax']
        challan.igst = request.POST['igst_tax']
        challan.tax_amount = request.POST['tax_amount']
        challan.adjustment = request.POST['adjustment']
        challan.total_amount = request.POST['grand_total']
        challan.balance = 0
        challan.status = 'Open'
        challan.is_converted = False

        challan.save()

        ids = request.POST.getlist('dcItems[]')
        item = request.POST.getlist("item[]")
        hsn  = request.POST.getlist("hsn[]")
        qty = request.POST.getlist("qty[]")
        price = request.POST.getlist("price[]")
        tax = request.POST.getlist("taxgst[]") if request.POST['state_supply'] == 'state' else request.POST.getlist("taxigst[]")
        discount = request.POST.getlist("discount[]")
        total = request.POST.getlist("total[]")
        dc_item_ids = request.POST.getlist("id[]")
        
        item_ids = [int(id) for id in dc_item_ids]

        
        dc_item = DeliveryChallanItems.objects.filter(cid = challan)
        object_ids = [obj.id for obj in dc_item]

        ids_to_delete = [obj_id for obj_id in object_ids if obj_id not in item_ids]

        DeliveryChallanItems.objects.filter(id__in=ids_to_delete).delete()
        
        count = DeliveryChallanItems.objects.filter(cid = challan, company = com).count()
        if len(ids)==len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total):
            try:
                mapped=zip(ids,item,hsn,qty,price,tax,total,discount,item_ids)
                mapped=list(mapped)
                
                for ele in mapped:
                    if int(len(item))>int(count):
                        if ele[8] == 0:
                            itemAdd= DeliveryChallanItems.objects.create(name = ele[1], hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7] ,cid = challan, staff = staff, company = com, item = ItemModel.objects.get(company = com, id = ele[0]))
                        else:
                            itemAdd = DeliveryChallanItems.objects.filter( id = ele[8],company = com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
                    else:
                        itemAdd = DeliveryChallanItems.objects.filter( id = ele[8],company=com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
            except Exception as e:
                    print(e)
                    mapped=zip(ids,item,hsn,qty,price,tax,total,discount,item_ids)
                    mapped=list(mapped)
                    
                    for ele in mapped:
                        created =DeliveryChallanItems.objects.filter(id=ele[8] ,company=com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))

        history = DeliveryChallanTransactionHistory(
          staff = staff,
          challan = challan,
          company = com,
          action = "Edit"
        )
        history.save()

        return redirect(viewChallan,id)
    except Exception as e:
      print(e)
      return redirect(editChallan, id)


def challanTransactionHistory(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    try:
      dc = DeliveryChallan.objects.get(company = com, id = id)
      history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc)
      context = {
        'staff':staff, 'company':com, 'allmodules':allmodules, 'history':history,
      }
      return render(request, 'staff/delivery_challan_transaction_history.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def importEstimateFromExcel(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)    
    
    current_datetime = timezone.now()
    dateToday =  current_datetime.date()

    if request.method == "POST" and 'excel_file' in request.FILES:
    
        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)

        # checking estimate sheet columns
        try:
          ws = wb["estimate"]
        except:
          print('sheet not found')
          messages.error(request,'`estimate` sheet not found.! Please check.')
          return redirect(estimate_quotation)

        try:
          ws = wb["items"]
        except:
          print('sheet not found')
          messages.error(request,'`items` sheet not found.! Please check.')
          return redirect(estimate_quotation)
        
        ws = wb["estimate"]
        estimate_columns = ['SLNO','DATE','NAME','STATE OF SUPPLY','DESCRIPTION','SUB TOTAL','IGST','CGST','SGST','TAX AMOUNT','ADJUSTMENT','GRAND TOTAL']
        estimate_sheet = [cell.value for cell in ws[1]]
        if estimate_sheet != estimate_columns:
          print('invalid sheet')
          messages.error(request,'`estimate` sheet column names or order is not in the required formate.! Please check.')
          return redirect(estimate_quotation)

        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,name,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal = row
          if slno is None or state_of_supply is None or taxamount is None or grandtotal is None:
            print('estimate == invalid data')
            messages.error(request,'`estimate` sheet entries missing required fields.! Please check.')
            return redirect(estimate_quotation)
        
        # checking items sheet columns
        ws = wb["items"]
        items_columns = ['ESTIMATE NO','NAME','HSN','QUANTITY','PRICE','TAX PERCENTAGE','DISCOUNT','TOTAL']
        items_sheet = [cell.value for cell in ws[1]]
        if items_sheet != items_columns:
          print('invalid sheet')
          messages.error(request,'`items` sheet column names or order is not in the required formate.! Please check.')
          return redirect(estimate_quotation)

        for row in ws.iter_rows(min_row=2, values_only=True):
          est_no,name,hsn,quantity,price,tax_percentage,discount,total = row
          if est_no is None or name is None or quantity is None or tax_percentage is None or total is None:
            print('items == invalid data')
            messages.error(request,'`items` sheet entries missing required fields.! Please check.')
            return redirect(estimate_quotation)
        
        # getting data from estimate sheet and create estimate.
        ws = wb['estimate']
        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,name,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal = row
          estNo = slno
          if slno is None:
            continue
          # Fetching last bill and assigning upcoming bill no as current + 1
          # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
          latest_bill = Estimate.objects.filter(company = com).order_by('-id').first()
          
          if latest_bill:
              last_number = int(latest_bill.ref_no)
              new_number = last_number + 1
          else:
              new_number = 1

          if DeletedEstimate.objects.filter(company = com).exists():
              deleted = DeletedEstimate.objects.get(company = com)
              
              if deleted:
                  while int(deleted.ref_no) >= new_number:
                      new_number+=1
          try:
            cntct = party.objects.get(company = com, party_name = name).contact
            adrs = party.objects.get(company = com, party_name = name).address
          except:
            pass

          if date is None:
            date = dateToday

          print(date,new_number,name,cntct,adrs,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal)

          estimate = Estimate(
              staff = staff,
              company = com,
              date = date,
              ref_no = new_number,
              party_name = name,
              contact = cntct,
              billing_address = adrs,
              state_of_supply = 'State' if str(state_of_supply).lower() == 'state' else 'Other State',
              description = description,
              subtotal = subtotal,
              cgst = cgst,
              sgst = sgst,
              igst = igst,
              tax_amount = taxamount,
              adjustment = adjustment,
              total_amount = grandtotal,
              balance = 0,
              status = 'Open',
              is_converted = False
          )
          estimate.save()

          # Transaction history
          history = EstimateTransactionHistory(
            staff = staff,
            estimate = estimate,
            company = com,
            action = "Create"
          )
          history.save()

          # Items for the estimate
          ws = wb['items']
          for row in ws.iter_rows(min_row=2, values_only=True):
            est_no,name,hsn,quantity,price,tax_percentage,discount,total = row
            if int(est_no) == int(estNo):
              print(row)
              if estimate.state_of_supply == 'State' and tax_percentage:
                tx = 'GST'+str(tax_percentage)+'['+str(tax_percentage)+'%]'
              elif estimate.state_of_supply == 'Other State' and tax_percentage:
                tx = 'IGST'+str(tax_percentage)+'['+str(tax_percentage)+'%]'
              if discount is None:
                discount=0
              if price is None:
                price=0
              try:
                itm = ItemModel.objects.get(company = com, item_name = name)
              except:
                pass
              Estimate_items.objects.create(staff = staff, eid = estimate, company = com, item = itm,name = name,hsn=hsn,quantity=int(quantity),price = float(price),tax=tx, discount = float(discount),total=float(total))
    messages.success(request, 'Data imported successfully.!')
    return redirect(estimate_quotation)
  

def importChallanFromExcel(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)    
    
    current_datetime = timezone.now()
    dateToday =  current_datetime.date()

    if request.method == "POST" and 'excel_file' in request.FILES:
    
        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)

        # checking challan sheet columns
        try:
          ws = wb["challan"]
        except:
          print('sheet not found')
          messages.error(request,'`challan` sheet not found.! Please check.')
          return redirect(delivery_challan)

        try:
          ws = wb["items"]
        except:
          print('sheet not found')
          messages.error(request,'`items` sheet not found.! Please check.')
          return redirect(delivery_challan)
        
        ws = wb["challan"]
        estimate_columns = ['SLNO','DATE','DUE DATE','NAME','STATE OF SUPPLY','DESCRIPTION','SUB TOTAL','IGST','CGST','SGST','TAX AMOUNT','ADJUSTMENT','GRAND TOTAL']
        estimate_sheet = [cell.value for cell in ws[1]]
        if estimate_sheet != estimate_columns:
          print('invalid sheet')
          messages.error(request,'`challan` sheet column names or order is not in the required formate.! Please check.')
          return redirect(delivery_challan)

        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,due_date,name,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal = row
          if slno is None or state_of_supply is None or taxamount is None or grandtotal is None:
            print('challan == invalid data')
            messages.error(request,'`challan` sheet entries missing required fields.! Please check.')
            return redirect(delivery_challan)
        
        # checking items sheet columns
        ws = wb["items"]
        items_columns = ['CHALLAN NO','NAME','HSN','QUANTITY','PRICE','TAX PERCENTAGE','DISCOUNT','TOTAL']
        items_sheet = [cell.value for cell in ws[1]]
        if items_sheet != items_columns:
          print('invalid sheet')
          messages.error(request,'`items` sheet column names or order is not in the required formate.! Please check.')
          return redirect(delivery_challan)

        for row in ws.iter_rows(min_row=2, values_only=True):
          chl_no,name,hsn,quantity,price,tax_percentage,discount,total = row
          if chl_no is None or name is None or quantity is None or tax_percentage is None or total is None:
            print('items == invalid data')
            messages.error(request,'`items` sheet entries missing required fields.! Please check.')
            return redirect(delivery_challan)
        
        # getting data from estimate sheet and create estimate.
        ws = wb['challan']
        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,due_date,name,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal = row
          dcNo = slno
          if slno is None:
            continue
          # Fetching last bill and assigning upcoming bill no as current + 1
          # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
          latest_bill = DeliveryChallan.objects.filter(company = com).order_by('-id').first()
          
          if latest_bill:
              last_number = int(latest_bill.challan_no)
              new_number = last_number + 1
          else:
              new_number = 1

          if DeletedDeliveryChallan.objects.filter(company = com).exists():
              deleted = DeletedDeliveryChallan.objects.get(company = com)
              
              if deleted:
                  while int(deleted.challan_no) >= new_number:
                      new_number+=1
          try:
            cntct = party.objects.get(company = com, party_name = name).contact
            adrs = party.objects.get(company = com, party_name = name).address
          except:
            pass

          if date is None:
            date = dateToday

          if due_date is None:
            due_date = dateToday

          print(date,due_date,new_number,name,cntct,adrs,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal)

          challan = DeliveryChallan(
              staff = staff,
              company = com,
              date = date,
              due_date = due_date,
              challan_no = new_number,
              party_name = name,
              contact = cntct,
              billing_address = adrs,
              state_of_supply = 'State' if str(state_of_supply).lower() == 'state' else 'Other State',
              description = description,
              subtotal = subtotal,
              cgst = cgst,
              sgst = sgst,
              igst = igst,
              tax_amount = taxamount,
              adjustment = adjustment,
              total_amount = grandtotal,
              balance = 0,
              status = 'Open',
              is_converted = False
          )
          challan.save()

          # Transaction history
          history = DeliveryChallanTransactionHistory(
            staff = staff,
            challan = challan,
            company = com,
            action = "Create"
          )
          history.save()

          # Items for the estimate
          ws = wb['items']
          for row in ws.iter_rows(min_row=2, values_only=True):
            chl_no,name,hsn,quantity,price,tax_percentage,discount,total = row
            if int(chl_no) == int(dcNo):
              print(row)
              if challan.state_of_supply == 'State' and tax_percentage:
                tx = 'GST'+str(tax_percentage)+'['+str(tax_percentage)+'%]'
              elif challan.state_of_supply == 'Other State' and tax_percentage:
                tx = 'IGST'+str(tax_percentage)+'['+str(tax_percentage)+'%]'
              if discount is None:
                discount=0
              if price is None:
                price=0
              try:
                itm = ItemModel.objects.get(company = com, item_name = name)
              except:
                pass
              DeliveryChallanItems.objects.create(staff = staff, cid = challan, company = com, item = itm,name = name,hsn=hsn,quantity=int(quantity),price = float(price),tax=tx, discount = float(discount),total=float(total))
    messages.success(request, 'Data imported successfully.!')
    return redirect(delivery_challan)



def downloadEstimateSampleImportFile(request):
    
    estimate_table_data = [['SLNO','DATE','NAME','STATE OF SUPPLY','DESCRIPTION','SUB TOTAL','IGST','CGST','SGST','TAX AMOUNT','ADJUSTMENT','GRAND TOTAL'], ['1', '2023-11-20', 'Alwin', 'State', 'Sample Description','1000','0','25','25','50','0','1050']]
    items_table_data = [['ESTIMATE NO', 'NAME','HSN','QUANTITY','PRICE','TAX PERCENTAGE','DISCOUNT','TOTAL'], ['1', 'Test Item 1','789987','1','1000','5','0','1000']]

    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = 'estimate'
    sheet2 = wb.create_sheet(title='items')

    # Populate the sheets with data
    for row in estimate_table_data:
        sheet1.append(row)

    for row in items_table_data:
        sheet2.append(row)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=estimate_sample_file.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


def downloadChallanSampleImportFile(request):
    
    challan_table_data = [['SLNO','DATE','DUE DATE','NAME','STATE OF SUPPLY','DESCRIPTION','SUB TOTAL','IGST','CGST','SGST','TAX AMOUNT','ADJUSTMENT','GRAND TOTAL'], ['1', '2023-11-20', '2023-11-20', 'Alwin', 'State', 'Sample Description','1000','0','25','25','50','0','1050']]
    items_table_data = [['CHALLAN NO', 'NAME','HSN','QUANTITY','PRICE','TAX PERCENTAGE','DISCOUNT','TOTAL'], ['1', 'Test Item 1','788987','1','1000','5','0','1000']]

    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = 'challan'
    sheet2 = wb.create_sheet(title='items')

    # Populate the sheets with data
    for row in challan_table_data:
        sheet1.append(row)

    for row in items_table_data:
        sheet2.append(row)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=challan_sample_file.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


def estimateBillPdf(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)  

    bill = Estimate.objects.get(company = com, id = id)
    items = Estimate_items.objects.filter(company = com, eid = bill)

    total = bill.total_amount
    words_total = num2words(total)
    
    context = {'staff':staff,'bill': bill, 'company': com,'items':items, 'total':words_total}
    
    template_path = 'staff/estimate_bill_pdf.html'
    fname = 'bill'+str(bill.ref_no)

    # return render(request, 'staff/estimate_bill_pdf.html',context)
    # Create a Django response object, and specify content_type as pdftemp_creditnote
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] =f'attachment; filename = Estimate_{fname}.pdf'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def challanBillPdf(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)  

    bill = DeliveryChallan.objects.get(company = com, id = id)
    items = DeliveryChallanItems.objects.filter(company = com, cid = bill)

    total = bill.total_amount
    words_total = num2words(total)
    
    context = {'staff':staff,'bill': bill, 'company': com,'items':items, 'total':words_total}
    
    template_path = 'staff/challan_bill_pdf.html'
    fname = 'bill'+str(bill.challan_no)

    # return render(request, 'staff/challan_bill_pdf.html',context)
    # Create a Django response object, and specify content_type as pdftemp_creditnote
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] =f'attachment; filename = DeliveryChallan_{fname}.pdf'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def viewEstimate(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    try:
      bill = Estimate.objects.get(company = com, id = id)
      items = Estimate_items.objects.filter(company = com , eid = bill)
      context= {
        'staff':staff, 'company':com, 'bill':bill, 'items': items,'allmodules':allmodules
      }
      return render(request, 'staff/view_estimate.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    

def viewChallan(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    try:
      bill = DeliveryChallan.objects.get(company = com, id = id)
      items = DeliveryChallanItems.objects.filter(company = com , cid = bill)
      context= {
        'staff':staff, 'company':com, 'bill':bill, 'items': items,'allmodules':allmodules
      }
      return render(request, 'staff/view_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
    

def addNewParty(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)

    if request.method == 'POST':
      Company = company.objects.get(id = staff.company.id)
      user_id = request.user.id
      
      party_name = request.POST['partyname']
      gst_no = request.POST['gstno']
      contact = request.POST['contact']
      gst_type = request.POST['gst']
      state = request.POST['state']
      address = request.POST['address']
      email = request.POST['email']
      openingbalance = request.POST.get('balance', '')
      payment = request.POST.get('paymentType', '')
      creditlimit = request.POST.get('creditlimit', '')
      current_date = request.POST['currentdate']
      End_date = request.POST.get('enddate', None)
      additionalfield1 = request.POST['additionalfield1']
      additionalfield2 = request.POST['additionalfield2']
      additionalfield3 = request.POST['additionalfield3']
      comp=Company
      if (not party_name):
        return render(request, 'add_parties.html')

      part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                      creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,company=comp)
      part.save()

      return JsonResponse({'status':True})


def addNewItem(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    if request.method=='POST':
      company_user_data = com
      item_name = request.POST.get('item_name')
      item_hsn = request.POST.get('item_hsn')
      item_unit = request.POST.get('item_unit')
      item_taxable = request.POST.get('item_taxable')
      item_gst = request.POST.get('item_gst')
      item_igst = request.POST.get('item_igst')
      item_sale_price = request.POST.get('item_sale_price')
      item_purchase_price = request.POST.get('item_purchase_price')
      item_opening_stock = request.POST.get('item_opening_stock')
      item_current_stock = item_opening_stock
      if item_opening_stock == '' or None :
        item_opening_stock = 0
        item_current_stock = 0
      item_at_price = request.POST.get('item_at_price')
      if item_at_price == '' or None:
        item_at_price =0
      item_date = request.POST.get('item_date')
      item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
      if item_min_stock_maintain == ''  or None:
        item_min_stock_maintain = 0
      item_data = ItemModel(company=company_user_data,
        item_name=item_name,
        item_hsn=item_hsn,
        item_unit=item_unit,
        item_taxable=item_taxable,
        item_gst=item_gst,
        item_igst=item_igst,
        item_sale_price=item_sale_price,
        item_purchase_price=item_purchase_price,
        item_opening_stock=item_opening_stock,
        item_current_stock=item_current_stock,
        item_at_price=item_at_price,
        item_date=item_date,
        item_min_stock_maintain=item_min_stock_maintain
      )
      item_data.save()

      return JsonResponse({'status':True})
    

# ===================end ---shemeem =============================


# ----------athul-22-11-2023--------

def register(request):
  if request.method == 'POST':
    first_name = request.POST['fname']
    last_name = request.POST['lname']
    user_name = request.POST['uname']
    email_id = request.POST['eid']
    mobile = request.POST['ph']
    passw = request.POST['pass']
    c_passw = request.POST['cpass']
    action = request.POST['r']
    did = request.POST['did']
    if did != '':
      if Distributors_details.objects.filter(distributor_id=did).exists():
        distributor = Distributors_details.objects.get(distributor_id=did)
      else :
          messages.info(request, 'Sorry, distributor id does not exists')
          return redirect('company_reg')
    

    
    if passw == c_passw:
      if User.objects.filter(username = user_name).exists():
        messages.info(request, 'Sorry, Username already exists')
        return redirect('company_reg')
      

      elif not User.objects.filter(email = email_id).exists():
        
        user_data = User.objects.create_user(first_name = first_name,
                        last_name = last_name,
                        username = user_name,
                        email = email_id,
                        password = passw)
        user_data.save()
        if did != '':
          data = User.objects.get(id = user_data.id)
          cust_data = company( contact=mobile,
                             user = data,reg_action=action,Distributors=distributor)
          cust_data.save()
          demo_staff=staff_details(company=cust_data,
                                   email=email_id,
                                   position='company',
                                   user_name=user_name,
                                   password=passw,
                                   contact=mobile)
          demo_staff.save()
          return redirect('company_reg2')
        else:
          data = User.objects.get(id = user_data.id)
          cust_data = company( contact=mobile,
                             user = data,reg_action=action)
          cust_data.save()
          demo_staff=staff_details(company=cust_data,
                                   email=email_id,
                                   position='company',
                                   user_name=user_name,
                                   password=passw,
                                   contact=mobile)
          demo_staff.save()

          print(demo_staff.company.company_name)
        
          return redirect('company_reg2')
      else:
        messages.info(request, 'Sorry, Email already exists')
        return redirect('company_reg')
    return render(request,'company/register.html')
  
def Allmodule(request,uid):
  user=User.objects.get(id=uid)
  return render(request,'company/modules.html',{'user':user})

def addmodules(request,uid):
  if request.method == 'POST':
    com=company.objects.get(user=uid)
    c1=request.POST.get('c1')
    c2=request.POST.get('c2')
    c3=request.POST.get('c3')
    c4=request.POST.get('c4')
    c5=request.POST.get('c5')
    c6=request.POST.get('c6')
    c7=request.POST.get('c7')
    c8=request.POST.get('c8')
    c9=request.POST.get('c9')
    c10=request.POST.get('c10')
    c11=request.POST.get('c11')
    c12=request.POST.get('c12')
    c13=request.POST.get('c13')
    c14=request.POST.get('c14')
    c15=request.POST.get('c15')
    
    data=modules_list(company=com,sales_invoice = c1,
                      Estimate=c2,Payment_in=c3,sales_order=c4,
                      Delivery_challan=c5,sales_return=c6,Purchase_bills=c7,
                      Payment_out=c8,Purchase_order=c9,Purchase_return=c10,
                      Bank_account=c11,Cash_in_hand=c12, cheques=c13,Loan_account=c14,Upi=c15)
    data.save()

    return redirect('log_page')
    
def adminaccept(request,id):
  data=company.objects.filter(id=id).update(superadmin_approval=1)
  data1=staff_details.objects.filter(company=id,position='company').update(Action=1)
  return redirect('client_request')
def adminreject(request,id):
  data1=staff_details.objects.get(company=id,position='company')
  data1.delete()
  data=company.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('client_request')


def log_page(request):
  return render(request, 'log.html')
  
def login(request):
  if request.method == 'POST':
    user_name = request.POST['username']
    passw = request.POST['password']
    
    log_user = auth.authenticate(username = user_name,
                                  password = passw)
    
    if log_user is not None:
      auth.login(request, log_user)

      # ---super admin---

      if request.user.is_staff==1:
        return redirect('adminhome')
      
      if Distributors_details.objects.filter(user=request.user).exists():
        data=Distributors_details.objects.get(user=request.user)
        if data.Log_Action == 1:
            return redirect('distributor_home')
        else:
            messages.info(request, 'Approval is Pending..')
            return redirect('log_page')
        
    if staff_details.objects.filter(user_name=user_name,password=passw,position='company').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='company') 

      if data.company.superadmin_approval == 1 or data.company.Distributor_approval == 1:
        request.session["staff_id"]=data.id
        if 'staff_id' in request.session:
          if request.session.has_key('staff_id'):
            staff_id = request.session['staff_id']
            print(staff_id)
 
          return redirect('homepage')  
      else:
        messages.info(request, 'Approval is Pending..')
        return redirect('log_page')
      
    if staff_details.objects.filter(user_name=user_name,password=passw,position='staff').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='staff')   
      if data.Action == 1:
        request.session["staff_id"]=data.id
        if 'staff_id' in request.session:
          if request.session.has_key('staff_id'):
            staff_id = request.session['staff_id']
            print(staff_id)
 
            return redirect('staffhome')  
      else:
        messages.info(request, 'Approval is Pending..')
        return redirect('log_page')
    else:
      messages.info(request, 'Invalid Username or Password. Try Again.')
      return redirect('log_page')  
  else:  
   return redirect('log_page')   
  

def homepage(request):
 
  staff_id = request.session['staff_id']
  print(staff_id)
       
  staff =  staff_details.objects.get(id = staff_id)
  print(staff.position)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules
          }
  return render(request, 'company/homepage.html', context)

def staff_request(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  data = staff_details.objects.filter(company=staff.company.id,Action=0,position='staff').order_by('-id')
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request,'company/staff_request.html',{'staff':staff,'data':data,'allmodules':allmodules}) 

# @login_required(login_url='login')
def staffhome(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  

  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request, 'staff/staffhome.html', context)


 
def View_staff(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  data = staff_details.objects.filter(company=staff.company.id,Action=1,position='staff').order_by('-id')
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')

  return render(request, 'company/view_staff.html',{'staff':staff,'data':data,'allmodules':allmodules})

def Companyprofile(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request,'company/companyprofile.html',{'staff':staff,'allmodules':allmodules})

def editcompanyprofile(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  terms=payment_terms.objects.all()
  return render(request,'company/editcompanyprofile.html',{'staff':staff,'allmodules':allmodules,'terms':terms})

def editcompanyprofile_action(request):
  staff_id = request.session['staff_id']
  print(staff_id) 
  staff =  staff_details.objects.get(id = staff_id)
 
  if request.method == 'POST':
    staff.company.company_name = request.POST['cname']
    staff.company.user.email = request.POST['email']

    staff.email = request.POST['email']

    staff.company.contact = request.POST['ph']

    staff.contact = request.POST['ph']

    staff.company.address = request.POST['address']
    staff.company.city = request.POST['city']
    staff.company.state = request.POST['state']
    staff.company.country = request.POST['country']
    staff.company.pincode = request.POST['pincode']

    t = request.POST['select']
    terms = payment_terms.objects.get(id=t)
    staff.company.dateperiod = terms
    staff.company.start_date=date.today()
    days=int(terms.days)

    end= date.today() + timedelta(days=days)
    staff.company.End_date=end

    old=staff.company.profile_pic
    new=request.FILES.get('image')
    if old!=None and new==None:
      staff.company.profile_pic=old
    else:
      staff.company.profile_pic=new
    
    staff.company.save() 
    staff.company.user.save() 
    staff.save()
    return redirect('Companyprofile') 



  return redirect('Companyprofile')


def editmodule(request):
  staff_id = request.session['staff_id']
  print(staff_id) 
  staff =  staff_details.objects.get(id = staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request,'company/editmodule.html',{'staff':staff,'allmodules':allmodules})

def editmodule_action(request):
  if request.method == 'POST':
    staff_id = request.session['staff_id']
    print(staff_id) 
    staff =  staff_details.objects.get(id = staff_id)
    com = company.objects.get(id = staff.company.id)
    # if modules_list.objects.filter(company=com.id,status='Old').exists():
    #   old=modules_list.objects.filter(company=com.id,status='Old')
    #   old.delete()

    # old_data=modules_list.objects.get(company=com.id,status='New')  
    # old_data.status='Old'
    # old_data.save()



    c1=request.POST.get('c1')
    c2=request.POST.get('c2')
    c3=request.POST.get('c3')
    c4=request.POST.get('c4')
    c5=request.POST.get('c5')
    c6=request.POST.get('c6')
    c7=request.POST.get('c7')
    c8=request.POST.get('c8')
    c9=request.POST.get('c9')
    c10=request.POST.get('c10')
    c11=request.POST.get('c11')
    c12=request.POST.get('c12')
    c13=request.POST.get('c13')
    c14=request.POST.get('c14')
    c15=request.POST.get('c15')
    
    data=modules_list(company=com,sales_invoice = c1,
                      Estimate=c2,Payment_in=c3,sales_order=c4,
                      Delivery_challan=c5,sales_return=c6,Purchase_bills=c7,
                      Payment_out=c8,Purchase_order=c9,Purchase_return=c10,
                      Bank_account=c11,Cash_in_hand=c12, cheques=c13,Loan_account=c14,Upi=c15,status='Pending')
    data.save()
    data1=modules_list.objects.filter(company=com.id,status='Pending').update(update_action=1)
    return redirect('Companyprofile')
    
    
  return redirect('Companyprofile')


def companyreport(request):
  staff_id = request.session['staff_id']
  print(staff_id) 
  staff =  staff_details.objects.get(id = staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request,'company/companyreport.html',{'staff':staff,'allmodules':allmodules}) 





def staff_profile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request,'staff/staff_profile.html',context)

def editstaff_profile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request,'staff/editstaff_profile.html',context)

def editstaff_profile_action(request):
  if request.method == 'POST':
    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    staff.first_name = request.POST['fname']
    staff.last_name = request.POST['lname']
    staff.user_name = request.POST['uname']
    staff.email = request.POST['email']
    staff.contact = request.POST['ph']
    old=staff.img
    new=request.FILES.get('image')
    if old!=None and new==None:
      staff.img=old
    else:
      staff.img=new

    staff.save()  

    return redirect ('staff_profile')
  return redirect ('staff_profile')

def view_parties(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  
 
  Party=party.objects.filter(company=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  return render(request, 'company/view_parties.html',{'staff':staff,'allmodules':allmodules,'Party':Party})

def save_parties(request):
    if request.method == 'POST':
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        
        party_name = request.POST['partyname']
        gst_no = request.POST['gstno']
        contact = request.POST['contact']
        gst_type = request.POST['gst']
        state = request.POST['state']
        address = request.POST['address']
        email = request.POST['email']
        openingbalance = request.POST.get('balance', '')
        payment = request.POST.get('paymentType', '')
        creditlimit = request.POST.get('creditlimit', '')
        current_date = request.POST['currentdate']
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST['additionalfield1']
        additionalfield2 = request.POST['additionalfield2']
        additionalfield3 = request.POST['additionalfield3']
       
        if (
          not party_name
          
      ):
          return render(request, 'add_parties.html')

        part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                       creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,user=staff.company.user,company=staff.company)
        part.save() 

        if 'save_and_new' in request.POST:
            
            return render(request, 'company/add_parties.html')
        else:
          
            return redirect('view_parties')

    return render(request, 'company/add_parties.html')  

def view_party(request,id):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  getparty=party.objects.get(id=id)
  Party=party.objects.filter(company=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  return render(request, 'company/view_party.html',{'staff':staff,'allmodules':allmodules,'Party':Party,'getparty':getparty})


#********************************************keerthana***********************************************************

 

def create_sale(request):
    toda = date.today()
    tod = toda.strftime("%Y-%m-%d")
   
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    Party = party.objects.filter(company=cmp,user=cmp.user)
    bank = BankModel.objects.filter(company=cmp,user=cmp.user)
    allmodules= modules_list.objects.get(company=staff.company,status='New')
    
    last_credit = CreditNote.objects.filter(company=cmp).count()

    if last_credit:
      credit_note = last_credit + 1 
    else:
      credit_note = 1

    item = ItemModel.objects.filter(company=cmp,user=cmp.user)
    item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company)
            
       

      
    context = {'staff':staff, 'allmodules':allmodules, 'party':Party, 'cmp':cmp,'credit_note':credit_note,'tod':tod,'item':item, 'item_units':item_units,'bank':bank}
    return render(request, 'company/create_sale.html', context)
# from datetime import datetime
# def add_creditnote(request):
#   if request.method == 'POST':
#     sid = request.session.get('staff_id')
#     staff = staff_details.objects.get(id=sid)
#     cmp = company.objects.get(id=staff.company.id)    
#     party_id = request.POST.get('partyname')
#     print(party_id)
#     part = party.objects.get(id=party_id)
#     return_no=request.POST.get('creditno')
#     partmob=request.POST.get('partyPhoneNumber')
#     creditdate=request.POST.get('cr_date')
#     invoiceno=request.POST.get('inv_no')
  
#     invoice_date=request.POST.get('inv_date')
    
#     supplyplace =request.POST.get('destination')
#     pay_method=request.POST.get("method")
#     cheque_no=request.POST.get("cheque_id")
#     upi_no=request.POST.get("upi_id")
#     bank_acc=request.POST.get("bnk_id")
#     advance = request.POST.get("advance")
#     balance = request.POST.get("balance")
#     subtotal=float(request.POST.get('subtotal'))
#     igst = request.POST.get('igst')
#     cgst = request.POST.get('cgst')
#     sgst = request.POST.get('sgst')
#     adjust = request.POST.get("adj")
#     taxamount = request.POST.get("taxamount")
#     grandtotal=request.POST.get('grandtotal')
#     descptn=request.POST.get('description')
#     creditnote=CreditNote(party=part,retrn_no = return_no,partymob=partmob,date=creditdate,invoice_date=invoice_date,
#                          invoiceno=invoiceno,supplyplace=supplyplace,pay_method=pay_method,cheque_no=cheque_no,upi_no=upi_no,
#                           bankaccount=bank_acc, subtotal=subtotal,advance=advance, balance=balance, igst=igst,
#                           cgst=cgst,sgst=sgst,adjust=adjust,taxamount=taxamount,grandtotal=grandtotal, description=descptn,
#                           company=cmp,staff=staff,)
    
#     #creditnote.save()
    
#     # product = request.POST["display"]
#     # display_ids = request.POST.get('display', '').split(',')
  
#     product =  tuple(request.POST.getlist("product[]"))
#     qty =  tuple(request.POST.getlist("qty[]"))
#     rate=tuple(request.POST.getlist("price[]"))
#     discount =  tuple(request.POST.getlist("discount[]"))
#     #if request.POST.get('destination') =='State':
#     tax =  tuple(request.POST.getlist("tax[]"))
#     # else:
#     #   tax =  tuple(request.POST.getlist("tax[]"))
#     total =  tuple(request.POST.getlist("total[]"))
#     return_no = CreditNote.objects.filter(retrn_no=creditnote.retrn_no, company=cmp).first()


#     if len(product)==len(qty)==len(tax)==len(discount)==len(total):
#       mapped=zip(product,qty,tax,discount,total)
#       mapped=list(mapped)
#       print('mapped',mapped)
#       for ele in mapped:
#         try:
#             item_id = int(ele[0])
#             itm = ItemModel.objects.get(id=item_id)
#         except (ValueError, ItemModel.DoesNotExist):
#             # Handle the case where 'ele[0]' is not a valid integer or the ItemModel does not exist
#             itm = None  # or another appropriate default value

#         if itm:
       
#           CreditNoteItem.objects.create(product=itm, qty=ele[1], tax=ele[2], discount=ele[3], total=ele[4], creditnote=return_no, company=cmp)


    
#     CreditNote.objects.filter(company=cmp).update(tot_credit_no=F('tot_credit_no') + 1)
    
#     creditnote.tot_credit_no = creditnote.retrn_no
#     creditnote.save()
#     CreditNoteTransactionHistory.objects.create(creditnote=creditnote,company=cmp,staff=staff,action='Created')
#     if 'save_and_new' in request.POST:
#       return redirect('create_sale')
#     if 'save' in request.POST:
#       return redirect('creditnote_list')
#   return render(request, 'company/create_sale.html')


def add_creditnote(request):
  if request.method == 'POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)    
    party_id = request.POST.get('partyname')
    print(party_id)
    part = party.objects.get(id=party_id)
    return_no=request.POST.get('creditno')
    partmob=request.POST.get('partyPhoneNumber')
    creditdate=request.POST.get('cr_date')
    invoiceno=request.POST.get('inv_no')
  
    invoice_date=request.POST.get('inv_date')
    supplyplace =request.POST.get('destination')
    pay_method=request.POST.get("method")
    cheque_no=request.POST.get("cheque_id")
    upi_no=request.POST.get("upi_id")
    bank_acc=request.POST.get("bnk_id")
    advance = request.POST.get("advance")
    balance = request.POST.get("balance")
    subtotal=float(request.POST.get('subtotal'))
    igst = request.POST.get('igst')
    cgst = request.POST.get('cgst')
    sgst = request.POST.get('sgst')
    adjust = request.POST.get("adj")
    taxamount = request.POST.get("taxamount")
    grandtotal=request.POST.get('grandtotal')
    descptn=request.POST.get('description')
    creditnote=CreditNote(party=part,retrn_no = return_no,partymob=partmob,date=creditdate,invoice_date=invoice_date,
                         invoiceno=invoiceno,supplyplace=supplyplace,pay_method=pay_method,cheque_no=cheque_no,upi_no=upi_no,
                          bankaccount=bank_acc, subtotal=subtotal,advance=advance, balance=balance, igst=igst,
                          cgst=cgst,sgst=sgst,adjust=adjust,taxamount=taxamount,grandtotal=grandtotal, description=descptn,
                          company=cmp,staff=staff,)
    
    creditnote.save()
    
    product = tuple(request.POST.getlist("product[]"))
    print('product',product)
    qty =  tuple(request.POST.getlist("qty[]"))
    rate=tuple(request.POST.getlist("price[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    
    tax =  tuple(request.POST.getlist("tax[]"))

    total =  tuple(request.POST.getlist("total[]"))
    return_no = CreditNote.objects.filter(retrn_no=creditnote.retrn_no, company=cmp).first()


    if len(product)==len(qty)==len(tax)==len(discount)==len(total):
      mapped=zip(product,qty,tax,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        CreditNoteItem.objects.create(product = itm,qty=ele[1], tax=ele[2],discount=ele[3],total=ele[4],creditnote=return_no,company=cmp)

    
    CreditNote.objects.filter(company=cmp).update(tot_credit_no=F('tot_credit_no') + 1)
    
    creditnote.tot_credit_no = creditnote.retrn_no
    creditnote.save()
    CreditNoteTransactionHistory.objects.create(creditnote=creditnote,company=cmp,staff=staff,action='Created')
    if 'save_and_new' in request.POST:
      return redirect('create_sale')
    if 'save' in request.POST:
      return redirect('creditnote_list')
  return render(request, 'company/create_sale.html')
    


  

def new_creditnote_item(request):
  print('items')
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  name = request.POST['name']
  unit = request.POST['unit']
  hsn = request.POST['hsn']
  taxref = request.POST['taxref']
  sell_price = request.POST['sell_price']
  cost_price = request.POST['cost_price']
  intra_st = request.POST['intra_st']
  inter_st = request.POST['inter_st']

  if taxref != 'Taxable':
    intra_st = 'GST0[0%]'
    inter_st = 'IGST0[0%]'

  itmdate = request.POST.get('itmdate')
  stock = request.POST.get('stock')
  itmprice = request.POST.get('itmprice')
  minstock = request.POST.get('minstock')

  if not hsn:
    hsn = None

  itm = ItemModel(item_name=name, item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
  itm.save() 
  return JsonResponse({'success': True})
  
from django.http import Http404
def get_hsn_for_item(request):
 
  itmid = request.GET.get('id')
  print(itmid)
  try:
      itm = ItemModel.objects.get(id=itmid)
      hsn = itm.item_hsn
      gst = itm.item_gst
      igst = itm.item_igst
      print(igst)
      price = itm.item_purchase_price
      qty = itm.item_current_stock
      return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})
  except ItemModel.DoesNotExist:
     raise Http404("Item not found")







def creditnote_list(request):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    allmodules = modules_list.objects.get(company=cmp, status='New')
    Part = party.objects.filter(company=cmp, user=cmp.user)
    credit = CreditNote.objects.filter(company=cmp)
       
    for i in credit:
        last_transaction = CreditNoteTransactionHistory.objects.filter(creditnote=i).last()
        if last_transaction:
            i.action = last_transaction.action
            
        else:
            i.action = None
            

    if not credit:
        context = {'staff': staff, 'allmodules': allmodules, 'party': Part}
        return render(request, 'company/salesfirst.html', context)


    context = {'staff': staff, 'allmodules': allmodules, 'credit': credit, 'party': Part,}
    return render(request, 'company/creditlist.html', context)







def party_dropdown(request):
    sid = request.session.get('staff_id')
    staff = get_object_or_404(staff_details, id=sid)
    cmp = get_object_or_404(company, id=staff.company.id)

    # Filter parties based on company and user
    part = party.objects.filter(company=cmp, user=cmp.user)

    # Extract party IDs and names
    id_list = [p.id for p in part]
    party_list = [p.party_name for p in part]

    # Return the data as JSON
    return JsonResponse({'id_list': id_list, 'party_list': party_list})



def saveparty(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  party_name = request.POST['name']
  email = request.POST['email']
  contact = request.POST['mobile']
  state = request.POST['splystate']
  address = request.POST['baddress']
  gst_type = request.POST['gsttype']
  gst_no = request.POST['gstin']
  current_date = request.POST['partydate']
  openingbalance = request.POST.get('openbalance')
  payment = request.POST.get('paytype')
  creditlimit = request.POST.get('credit_limit')
  End_date = request.POST.get('enddate', None)
  additionalfield1 = request.POST['add1']
  additionalfield2 = request.POST['add2']
  additionalfield3 = request.POST['add3']

  part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,
                payment=payment,creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,
                additionalfield3=additionalfield3,company=cmp,user=cmp.user)
  part.save() 
  return JsonResponse({'success': True})



def credit_bankdetails(request):
  bid = request.POST['id']
  bank = BankModel.objects.filter(id=bid) 
  bank_no = bank.account_num 
  bank_name = bank.bank_name
  return JsonResponse({'bank_no':bank_no,'bank_name':bank_name})



def detail_creditnote(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  allmodules = modules_list.objects.get(company=staff.company,status='New')
  credit = CreditNote.objects.get(id=id,company=cmp)
  citm = CreditNoteItem.objects.filter(creditnote=credit,company=cmp)
  dis = 0
  for itm in citm:
    dis += int(itm.discount)
  itm_len = len(citm)

  context={'staff':staff,'allmodules':allmodules,'credit':credit,'citm':citm,'itm_len':itm_len,'dis':dis}
  return render(request,'company/creditnotedetails.html',context)



# def creditnotehistory(request):
#   pid = request.POST['id']
#   sid = request.session.get('staff_id')
#   staff = staff_details.objects.get(id=sid)
#   cmp = company.objects.get(id=staff.company.id) 
#   credit = CreditNote.objects.get(no=pid,company=cmp)
#   hst = CreditNoteTransactionHistory.objects.filter(creditnote=credit,company=cmp).last()
#   name = hst.staff.first_name + ' ' + hst.staff.last_name 
#   action = hst.action
#   return JsonResponse({'name':name,'action':action,'pid':pid})

def import_creditnote(request):
  if request.method == 'POST' and request.FILES['creditfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(CreditNote.objects.filter(company=cmp).last().tot_bill_no) + 1

    excel_bill = request.FILES['creditfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=billsheet[0],email=billsheet[1],company=cmp)
      CreditNote.objects.create(party=part,retrn_no=totval,
                                  date=billsheet[2],
                                  supplyplace =billsheet[3],
                                  tot_bill_no = totval,
                                  company=cmp,staff=staff)
      
      credit = CreditNote.objects.last()
      if billsheet[4] == 'Cheque':
        credit.pay_method = 'Cheque'
        credit.cheque_no = billsheet[5]
      elif billsheet[4] == 'UPI':
        credit.pay_method = 'UPI'
        credit.upi_no = billsheet[5]
      else:
        if billsheet[4] != 'Cash':
          bank = BankModel.objects.get(bank_name=billsheet[4],company=cmp)
          credit.pay_method = bank
        else:
          credit.pay_method = 'Cash'
      credit.save()

      CreditNote.objects.filter(company=cmp).update(tot_bill_no=totval)
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=prdsheet[2])
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[5])
      CreditNoteItem.objects.create(creditnote=credit,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                tax=prdsheet[4],
                                discount=prdsheet[5],
                                total=total)

      temp = prdsheet[4].split('[')
      if billsheet[3] =='State':
        tax=int(temp[0][3:])
      else:
        tax=int(temp[0][4:])

        subtotal += total
        tamount = total *(tax / 100)
        taxamount += tamount
                
      if billsheet[3]=='State':
        gst = round((taxamount/2),2)
        credit.sgst=gst
        credit.cgst=gst
        credit.igst=0

      else:
        gst=round(taxamount,2)
        credit.igst=gst
        credit.cgst=0
        credit.sgst=0

      gtotal = subtotal + taxamount + float(billsheet[6])
      balance = gtotal- float(billsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      credit.subtotal=round(subtotal,2)
      credit.taxamount=round(taxamount,2)
      credit.adjust=round(billsheet[6],2)
      credit.grandtotal=gtotal
      credit.advance=round(billsheet[7],2)
      credit.balance=balance
      credit.save()

      CreditNoteTransactionHistory.objects.create(creditnote=credit,staff=credit.staff,company=credit.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})


def delete_CreditNote(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  credit = CreditNote.objects.get(id=id)
  CreditNoteItem.objects.filter(creditnote=credit,company=cmp).delete()
  credit.delete()
  return redirect('creditnote_list')

def edit_creditnote(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part= party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(user=cmp.user,company=staff.company.id)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  credit = CreditNote.objects.get(id=id,company=cmp)
  credit_itm = CreditNoteItem.objects.filter(creditnote=credit,company=cmp)


  if credit.pay_method != 'Cash' and credit.pay_method != 'Cheque' and credit.pay_method != 'UPI':
    try:
     bankno = BankModel.objects.get(id=credit.pay_method,company=cmp,user=cmp.user)
    except BankModel.DoesNotExist:
      bankno = 0
  else:
    bankno = 0

  bdate = credit.date.strftime("%Y-%m-%d")
  context = {'staff':staff, 'allmodules':allmodules, 'credit':credit, 'credititm':credit_itm,'tod':tod,
             'party':part, 'product':item, 'item_units':item_units, 'bdate':bdate,'bank':bank, 'bankno':bankno}
  return render(request,'company/edit_creditnot.html',context)


def update_creditnote(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    part = party.objects.get(id=request.POST.get('partyname'))
    credit = CreditNote.objects.get(id=id,company=cmp)
    credit.party = part
    credit.date = request.POST.get('cr_date')
    credit.supplyplace  = request.POST.get('destination')
    credit.subtotal =float(request.POST.get('subtotal'))
    credit.grandtotal = request.POST.get('grandtotal')
    credit.igst = request.POST.get('igst')
    credit.cgst = request.POST.get('cgst')
    credit.sgst = request.POST.get('sgst')
    credit.taxamount = request.POST.get("taxamount")
    credit.adjust = request.POST.get("adj")
    credit.pay_method = request.POST.get("method")
    credit.cheque_no = request.POST.get("cheque_id")
    credit.upi_no = request.POST.get("upi_id")
    credit.advance = request.POST.get("advance")
    credit.balance = request.POST.get("balance")

    credit.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    
    tax =tuple( request.POST.getlist("tax[]"))
      
  
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    CreditNoteItem.objects.filter(creditnote=credit,company=cmp).delete()
    if len(total)==len(discount)==len(qty)==len(tax):
      mapped=zip(product,qty,tax,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        CreditNoteItem.objects.create(product =itm,qty=ele[1], tax=ele[2],discount=ele[3],total=ele[4],creditnote=credit,company=cmp)

    CreditNoteTransactionHistory.objects.create(creditnote=credit,company=cmp,staff=staff,action='Updated')
    return redirect('creditnote_list')

  return redirect('creditnote_list')


def salesinvoicedata(request):
      try:
        selected_party_id = request.POST.get('id')
        print('id',selected_party_id)
        party_instance = get_object_or_404(party, id=selected_party_id)
        print('instance',party_instance)
        phone_number = party_instance.contact
        party_id = party_instance.id
        balnce =party_instance.openingbalance
    
    
        # Initialize lists to store multiple invoice numbers and dates
        invoice_numbers = []
        invoice_dates = []

        # Retrieve all SalesInvoice instances for the party
        invoice_instances = SalesInvoice.objects.filter(party=party_instance)
        print('invoice_instances',invoice_instances)

        # Loop through each SalesInvoice instance and collect invoice numbers and dates
        for invoice_instance in invoice_instances:
            invoice_numbers.append(invoice_instance.invoice_no)
            invoice_dates.append(invoice_instance.date)  # Format date as needed

        # Return a JSON response with the list of invoice numbers and dates
        if not invoice_numbers and not invoice_dates:
            return JsonResponse({'invoice_numbers': ['noinvoice'], 'invoice_dates': ['nodate'],'phone': phone_number, 'id': party_id,'balance':balnce})

        return JsonResponse({'invoice_numbers': invoice_numbers, 'invoice_dates': invoice_dates,'phone': phone_number, 'id': party_id,'balance':balnce})

      except party.DoesNotExist:
        return JsonResponse({'error': 'Party not found'}) 

@require_POST
@csrf_exempt
def get_Invoice_date(request):
    selected_inv_no = request.POST.get('inv_no', None)

    try:
        # Get the latest PurchaseBill with the specified bill_number
        salesinvoice = SalesInvoice.objects.filter(invoice_no=selected_inv_no).latest('inv_date')
        inv_date = salesinvoice.date.strftime('%Y-%m-%d')
    except SalesInvoice.DoesNotExist:
        return JsonResponse({'error': 'invoice not found'}, status=400)
    except SalesInvoice.MultipleObjectsReturned:
        # Handle the case where multiple PurchaseBills are found for the same bill_number
        return JsonResponse({'error': 'Multiple PurchaseBills found for the same bill number'}, status=400)

    return JsonResponse({'inv_date': inv_date})

    

def  creditnote_item_unit(request):
  if request.method=='POST':
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_unit_name = request.POST.get('item_unit_name')
    unit_data = UnitModel(user=user,company=company_user_data,unit_name=item_unit_name)
    unit_data.save()
  return JsonResponse({'message':'asdasd'})


def history_creditnote(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  credit = CreditNote.objects.get(id=id,company=cmp)
  hst= CreditNoteTransactionHistory.objects.filter(creditnote=credit,company=cmp)

  context = {'staff':staff,'allmodules':allmodules,'hst':hst,'credit':credit}
  return render(request,'company/creditnotehistory.html',context)


def  credititemdetails(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})


from django.views.decorators.cache import cache_control

import json
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def creditnote_item_dropdown(request):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    
    # Fetch products for the given company
    products = ItemModel.objects.filter(company=cmp, user=cmp.user)
    
    # Use a list comprehension to extract IDs and names
    id_list = [p.id for p in products]
    product_list = [p.item_name for p in products]
    response_data = {'id_list': id_list, 'product_list': product_list}
    print('Server response data:', json.dumps(response_data))

    return JsonResponse({'id_list': id_list, 'product_list': product_list})




def sharecreditnoteToEmail(request,id):
  if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

    
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)

                sid = request.session.get('staff_id')
                staff =  staff_details.objects.get(id=sid)
                cmp = company.objects.get(id=staff.company.id) 
               
                credit = CreditNote.objects.get(id=id,company=cmp)
                creditnoteitem = CreditNoteItem.objects.filter(creditnote=credit,company=cmp)
                        
                context = {'credit':credit, 'cmp':cmp,'creditnoteitem':creditnoteitem}
                template_path = 'company/creditnoteshare.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'CREDITNOTE - {credit.retrn_no}.pdf'
                subject = f"CREDITNOTE - {credit.retrn_no}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached INVOICE - File-{credit.retrn_no}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Invoice file has been shared via email successfully..!')
                return redirect(detail_creditnote,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(detail_creditnote, id)








    
